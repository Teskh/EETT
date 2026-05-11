from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.export_projection import iter_cost_model_rows, iter_material_context_rows


HEADER_FILL = PatternFill(fill_type="solid", start_color="D7E3F4", end_color="D7E3F4")
CATEGORY_FILL = PatternFill(fill_type="solid", start_color="E7EDF5", end_color="E7EDF5")
INSTANCE_FILL = PatternFill(fill_type="solid", start_color="F4F7FB", end_color="F4F7FB")
INPUT_FILL = PatternFill(fill_type="solid", start_color="FFF7D6", end_color="FFF7D6")
FORMULA_FILL = PatternFill(fill_type="solid", start_color="F3F4F6", end_color="F3F4F6")


def build_materials_workbook(project_data: dict[str, Any], output_path: Any) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    context_rows = list(iter_material_context_rows(project_data))

    totals_sheet = workbook.create_sheet("Total Materiales")
    _populate_total_materials_sheet(totals_sheet, context_rows)

    context_sheet = workbook.create_sheet("Por Contexto")
    _populate_context_sheet(context_sheet, context_rows, quantity_key="quantity", title="Q fabrica por categoria e instancia")

    assembly_sheet = workbook.create_sheet("Q obra")
    _populate_context_sheet(
        assembly_sheet,
        context_rows,
        quantity_key="assembly_quantity",
        title="Q obra por categoria e instancia",
    )

    workbook.properties.title = f"{project_data['project']['name']} - Libro de Materiales"
    if isinstance(output_path, Path):
        output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_cost_model_workbook(
    project_data: dict[str, Any],
    output_path: Any,
    *,
    prices_by_sku: dict[str, float | None],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    cost_rows = _build_cost_model_rows(project_data, prices_by_sku=prices_by_sku)

    instance_sheet = workbook.create_sheet("Por Instancia")
    _populate_cost_model_instance_sheet(instance_sheet, cost_rows, subtype_names=_flatten_subtype_names(project_data.get("subtypes", [])))

    totals_sheet = workbook.create_sheet("Total Materiales")
    _populate_cost_model_totals_sheet(totals_sheet, cost_rows, subtype_names=_flatten_subtype_names(project_data.get("subtypes", [])))

    workbook.properties.title = f"{project_data['project']['name']} - Modelo de Costos"
    if isinstance(output_path, Path):
        output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

def _populate_total_materials_sheet(ws, context_rows: list[dict[str, Any]]) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    include_subtype = any(
        (row.get("subtype") or "General") != "General"
        and _include_context_row(row, quantity_key="quantity", quantity_state_key="quantity_state")
        for row in context_rows
    )
    headers = ["Material", "SKU"]
    if include_subtype:
        headers.append("Subtipo")
    headers.extend(["Q fabrica", "Unidad"])
    ws.append(headers)
    _style_header_row(ws, row_index=1, column_count=len(headers))

    totals: dict[tuple[str, str], dict[str, Any]] = {}
    for row in context_rows:
        sku = row["sku"]
        subtype = row.get("subtype") or "General"
        key = (sku, subtype)
        entry = totals.setdefault(
            key,
            {
                "material_name": row["material_name"],
                "sku": sku,
                "subtype": subtype,
                "unit": row["unit"],
                "quantity_total": 0.0,
                "has_numeric_quantity": False,
                "has_blank_quantity": False,
            },
        )
        state = row["quantity_state"]
        value = row["quantity"]
        if state == "value" and value is not None:
            entry["has_numeric_quantity"] = True
            entry["quantity_total"] += float(value)
        elif state == "blank":
            entry["has_blank_quantity"] = True

    next_row = 2
    quantity_column = 4 if include_subtype else 3
    for entry in sorted(totals.values(), key=lambda item: (item["material_name"], item["sku"], item["subtype"] != "General", item["subtype"].lower())):
        if not entry["has_numeric_quantity"] and not entry["has_blank_quantity"]:
            continue
        quantity_value = entry["quantity_total"] if entry["has_numeric_quantity"] else None
        row_values = [entry["material_name"], entry["sku"]]
        if include_subtype:
            row_values.append(entry["subtype"])
        row_values.extend([quantity_value, entry["unit"]])
        ws.append(row_values)
        _style_data_row(ws, next_row, numeric_columns={quantity_column})
        next_row += 1

    _set_column_widths(ws, {"A": 42, "B": 16, "C": 18, "D": 14, "E": 10} if include_subtype else {"A": 42, "B": 16, "C": 14, "D": 10})


def _build_cost_model_rows(
    project_data: dict[str, Any],
    *,
    prices_by_sku: dict[str, float | None],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for row in iter_cost_model_rows(project_data):
        if row.get("quantity_state") == "zero":
            continue
        quantity_value = row.get("quantity") if row.get("quantity_state") == "value" else None
        rows.append(
            {
                "instance_name": row["instance_label"],
                "category_name": row["category_label"],
                "subtype_name": row["subtype"] or "General",
                "material_name": row["material_name"],
                "sku": row["sku"],
                "unit": row["unit"],
                "quantity": quantity_value,
                "price": _coerce_optional_float(prices_by_sku.get(str(row["sku"]).strip().upper())),
                "is_auxiliary": False,
            }
        )

    for auxiliary in sorted(
        project_data.get("auxiliary_materials", []),
        key=lambda item: ((item.get("code") or ""), (item.get("name") or "")),
    ):
        code = str(auxiliary.get("code") or "").strip().upper()
        if not code:
            continue
        rows.append(
            {
                "instance_name": "Materiales auxiliares",
                "category_name": auxiliary.get("category") or "Materiales auxiliares",
                "subtype_name": auxiliary.get("subtype") or "General",
                "material_name": auxiliary.get("name") or code,
                "sku": code,
                "unit": "",
                "quantity": 1.0,
                "price": _coerce_optional_float(auxiliary.get("price")),
                "is_auxiliary": True,
            }
        )

    return rows


def _populate_cost_model_instance_sheet(ws, cost_rows: list[dict[str, Any]], *, subtype_names: list[str]) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["Instancia", "Categoria", "Subtipo", "Material", "SKU", "Unidad", "Q fabrica", "Precio unitario", "Costo"]
    ws.append(headers)
    _style_header_row(ws, row_index=1, column_count=len(headers))
    _set_column_widths(ws, {"A": 28, "B": 24, "C": 18, "D": 40, "E": 16, "F": 10, "G": 14, "H": 14, "I": 14})

    if not cost_rows:
        ws.append(["No hay filas de costo disponibles para esta exportacion."])
        ws.merge_cells("A2:I2")
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        return

    currency_fmt = '$#,##0.00'
    quantity_fmt = "0.######"
    current_instance: str | None = None
    block_start_row: int | None = None

    for row in cost_rows:
        if current_instance is not None and row["instance_name"] != current_instance and block_start_row is not None:
            _append_instance_subtotal(ws, instance_name=current_instance, block_start_row=block_start_row, block_end_row=ws.max_row)
            current_instance = None
            block_start_row = None

        row_index = ws.max_row + 1
        ws.cell(row=row_index, column=1, value=row["instance_name"])
        ws.cell(row=row_index, column=2, value=row["category_name"])
        ws.cell(row=row_index, column=3, value=row["subtype_name"])
        ws.cell(row=row_index, column=4, value=row["material_name"])
        ws.cell(row=row_index, column=5, value=row["sku"])
        ws.cell(row=row_index, column=6, value=row["unit"])
        ws.cell(row=row_index, column=7, value=row["quantity"])
        ws.cell(row=row_index, column=8, value=row["price"])
        ws.cell(row=row_index, column=9, value=f'=IF(OR(G{row_index}="",H{row_index}=""),"",G{row_index}*H{row_index})')

        if current_instance is None:
            current_instance = row["instance_name"]
            block_start_row = row_index

        _style_data_row(ws, row_index, numeric_columns={7, 8, 9})
        ws.cell(row=row_index, column=7).number_format = quantity_fmt
        ws.cell(row=row_index, column=8).number_format = currency_fmt
        ws.cell(row=row_index, column=9).number_format = currency_fmt
        ws.cell(row=row_index, column=8).fill = INPUT_FILL
        ws.cell(row=row_index, column=9).fill = FORMULA_FILL

    if block_start_row is not None and current_instance is not None:
        _append_instance_subtotal(ws, instance_name=current_instance, block_start_row=block_start_row, block_end_row=ws.max_row)

    data_end_row = ws.max_row
    if subtype_names:
        for subtype_name in subtype_names:
            safe_name = subtype_name.replace('"', '""')
            row_index = ws.max_row + 1
            ws.append(["", "", "", f"Subtotal {subtype_name}", "", "", "", "", ""])
            ws.cell(
                row=row_index,
                column=9,
                value=f'=SUMIFS(I2:I{data_end_row},C2:C{data_end_row},"{safe_name}")+SUMIFS(I2:I{data_end_row},C2:C{data_end_row},"General")',
            )
            _style_cost_model_total_row(ws, row_index, label_column=4, value_column=9, currency_format=currency_fmt)

    last_row_before_total = ws.max_row
    total_row_index = last_row_before_total + 1
    ws.append(["", "", "", "Total general", "", "", "", "", ""])
    ws.cell(
        row=total_row_index,
        column=9,
        value=f'=SUMIFS(I2:I{last_row_before_total},C2:C{last_row_before_total},"General")',
    )
    _style_cost_model_total_row(ws, total_row_index, label_column=4, value_column=9, currency_format=currency_fmt)


def _append_instance_subtotal(ws, *, instance_name: str, block_start_row: int, block_end_row: int) -> None:
    row_index = ws.max_row + 1
    ws.append(["", "", "", f"Subtotal {instance_name}", "", "", "", "", ""])
    ws.cell(row=row_index, column=9, value=f"=SUM(I{block_start_row}:I{block_end_row})")
    _style_cost_model_total_row(ws, row_index, label_column=4, value_column=9, currency_format='$#,##0.00')


def _populate_cost_model_totals_sheet(ws, cost_rows: list[dict[str, Any]], *, subtype_names: list[str]) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["SKU", "Material", "Subtipo", "Unidad", "Q fabrica", "Precio unitario", "Costo"]
    ws.append(headers)
    _style_header_row(ws, row_index=1, column_count=len(headers))
    _set_column_widths(ws, {"A": 16, "B": 40, "C": 18, "D": 10, "E": 14, "F": 14, "G": 14})

    if not cost_rows:
        ws.append(["No hay filas de costo disponibles para esta exportacion."])
        ws.merge_cells("A2:G2")
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        return

    unique_keys: list[tuple[str, str, str, str]] = []
    seen_keys: set[tuple[str, str, str, str]] = set()
    for row in cost_rows:
        key = (row["sku"], row["material_name"], row["subtype_name"], row["unit"])
        if key in seen_keys:
            continue
        seen_keys.add(key)
        unique_keys.append(key)
    unique_keys.sort(key=lambda item: (item[0], item[1], item[2], item[3]))

    quantity_fmt = "0.######"
    currency_fmt = '$#,##0.00'
    for sku, material_name, subtype_name, unit in unique_keys:
        row_index = ws.max_row + 1
        ws.cell(row=row_index, column=1, value=sku)
        ws.cell(row=row_index, column=2, value=material_name)
        ws.cell(row=row_index, column=3, value=subtype_name)
        ws.cell(row=row_index, column=4, value=unit)
        sum_qty_formula = (
            f"SUMIFS('Por Instancia'!$G:$G,"
            f"'Por Instancia'!$E:$E,A{row_index},"
            f"'Por Instancia'!$C:$C,C{row_index})"
        )
        sum_cost_formula = (
            f"SUMIFS('Por Instancia'!$I:$I,"
            f"'Por Instancia'!$E:$E,A{row_index},"
            f"'Por Instancia'!$C:$C,C{row_index})"
        )
        ws.cell(row=row_index, column=5, value=f'=IF({sum_qty_formula}=0,"",{sum_qty_formula})')
        ws.cell(row=row_index, column=6, value=f'=IF(OR(E{row_index}="",E{row_index}=0,G{row_index}=""),"",G{row_index}/E{row_index})')
        ws.cell(row=row_index, column=7, value=f'=IF({sum_cost_formula}=0,"",{sum_cost_formula})')
        _style_data_row(ws, row_index, numeric_columns={5, 6, 7})
        ws.cell(row=row_index, column=5).number_format = quantity_fmt
        ws.cell(row=row_index, column=6).number_format = currency_fmt
        ws.cell(row=row_index, column=7).number_format = currency_fmt
        ws.cell(row=row_index, column=6).fill = FORMULA_FILL
        ws.cell(row=row_index, column=7).fill = FORMULA_FILL

    data_end_row = ws.max_row
    if subtype_names:
        for subtype_name in subtype_names:
            safe_name = subtype_name.replace('"', '""')
            row_index = ws.max_row + 1
            ws.append(["", "", f"Subtotal {subtype_name}", "", "", "", ""])
            ws.cell(
                row=row_index,
                column=7,
                value=f'=SUMIFS(G2:G{data_end_row},C2:C{data_end_row},"{safe_name}")+SUMIFS(G2:G{data_end_row},C2:C{data_end_row},"General")',
            )
            _style_cost_model_total_row(ws, row_index, label_column=3, value_column=7, currency_format=currency_fmt)

    total_row_index = ws.max_row + 1
    ws.append(["", "", "Total general", "", "", "", f'=SUMIFS(G2:G{data_end_row},C2:C{data_end_row},"General")'])
    _style_cost_model_total_row(ws, total_row_index, label_column=3, value_column=7, currency_format=currency_fmt)


def _populate_context_sheet(ws, context_rows: list[dict[str, Any]], *, quantity_key: str, title: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.append([title])
    title_cell = ws["A1"]
    title_cell.font = Font(name="Calibri", size=12, bold=True)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    quantity_state_key = f"{quantity_key}_state"
    visible_rows = [row for row in context_rows if _include_context_row(row, quantity_key=quantity_key, quantity_state_key=quantity_state_key)]
    include_subtype = any((row.get("subtype") or "General") != "General" for row in visible_rows)
    headers = ["Material", "SKU"]
    if include_subtype:
        headers.append("Subtipo")
    headers.extend(["Q fabrica" if quantity_key == "quantity" else "Q obra", "Unidad"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.append(headers)
    _style_header_row(ws, row_index=2, column_count=len(headers))

    if not visible_rows:
        ws.append(["No hay filas disponibles para esta exportacion."])
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
        ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
        _set_column_widths(ws, {"A": 42, "B": 16, "C": 18, "D": 14, "E": 10} if include_subtype else {"A": 42, "B": 16, "C": 14, "D": 10})
        return

    current_row = 3
    active_category: str | None = None
    active_instance: str | None = None

    for row in visible_rows:
        if row["category_label"] != active_category:
            ws.append([row["category_label"]])
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            _style_group_row(ws, current_row, fill=CATEGORY_FILL, bold=True)
            active_category = row["category_label"]
            active_instance = None
            current_row += 1

        if row["instance_label"] != active_instance:
            ws.append([f"  {row['instance_label']}"])
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            _style_group_row(ws, current_row, fill=INSTANCE_FILL, bold=False)
            active_instance = row["instance_label"]
            current_row += 1

        numeric_value = row[quantity_key] if row[quantity_state_key] == "value" else None
        row_values = [row["material_name"], row["sku"]]
        if include_subtype:
            row_values.append(row["subtype"] if row["subtype"] != "General" else "General")
        row_values.extend([numeric_value, row["unit"]])
        ws.append(row_values)
        _style_data_row(ws, current_row, numeric_columns={4 if include_subtype else 3})
        current_row += 1

    _set_column_widths(ws, {"A": 42, "B": 16, "C": 18, "D": 14, "E": 10} if include_subtype else {"A": 42, "B": 16, "C": 14, "D": 10})


def _include_context_row(row: dict[str, Any], *, quantity_key: str, quantity_state_key: str) -> bool:
    if quantity_key == "assembly_quantity":
        return row.get(quantity_state_key) == "value"
    state = row.get(quantity_state_key)
    return state in {"value", "blank"}


def _style_header_row(ws, *, row_index: int, column_count: int) -> None:
    for column_index in range(1, column_count + 1):
        cell = ws.cell(row=row_index, column=column_index)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_group_row(ws, row_index: int, *, fill: PatternFill, bold: bool) -> None:
    cell = ws.cell(row=row_index, column=1)
    cell.font = Font(name="Calibri", size=11, bold=bold)
    cell.fill = fill
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_data_row(ws, row_index: int, *, numeric_columns: set[int]) -> None:
    for column_index in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_index, column=column_index)
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(
            horizontal="right" if column_index in numeric_columns else "left",
            vertical="center",
            wrap_text=column_index == 1,
        )
        if column_index in numeric_columns and isinstance(cell.value, (int, float)):
            cell.number_format = "0.######"


def _style_cost_model_total_row(ws, row_index: int, *, label_column: int, value_column: int, currency_format: str) -> None:
    ws.cell(row=row_index, column=label_column).font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=row_index, column=label_column).fill = CATEGORY_FILL
    ws.cell(row=row_index, column=value_column).font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=row_index, column=value_column).fill = CATEGORY_FILL
    ws.cell(row=row_index, column=value_column).number_format = currency_format
    ws.cell(row=row_index, column=value_column).alignment = Alignment(horizontal="right", vertical="center")


def _set_column_widths(ws, widths: dict[str, float]) -> None:
    for column_letter, width in widths.items():
        ws.column_dimensions[column_letter].width = width


def _flatten_subtype_names(nodes: list[dict[str, Any]]) -> list[str]:
    names: list[str] = []
    for node in nodes:
        name = str(node.get("name") or "").strip()
        if name:
            names.append(name)
        names.extend(_flatten_subtype_names(node.get("children", [])))
    return names


def _coerce_optional_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None
