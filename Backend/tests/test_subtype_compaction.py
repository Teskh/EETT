from __future__ import annotations

from io import BytesIO
import unittest

from openpyxl import load_workbook

from app.services.export_projection import _technical_materials, build_detailed_material_export_sections
from app.services.export_workbooks import build_materials_workbook
from app.services.subtype_compaction import compact_subtype_rows


SUBTYPES = [
    {
        "id": 1,
        "name": "Base",
        "path": "Base",
        "kind": "variant",
        "children": [
            {
                "id": 10,
                "name": "Organization",
                "path": "Base › Organization",
                "kind": "group",
                "children": [
                    {
                        "id": 2,
                        "name": "Child",
                        "path": "Base › Organization › Child",
                        "kind": "variant",
                        "children": [
                            {
                                "id": 3,
                                "name": "Grandchild",
                                "path": "Base › Organization › Child › Grandchild",
                                "kind": "variant",
                                "children": [],
                            }
                        ],
                    }
                ],
            }
        ],
    }
]


def bom_row(subtype_id: int, quantity: float | None, state: str | None = None) -> dict:
    quantity_state = state or ("blank" if quantity is None else "zero" if quantity == 0 else "value")
    names = {
        1: "Base",
        2: "Base › Organization › Child",
        3: "Base › Organization › Child › Grandchild",
    }
    return {
        "subtype_id": subtype_id,
        "subtype": names[subtype_id],
        "quantity": quantity,
        "quantity_state": quantity_state,
        "effective_quantity": quantity,
        "effective_quantity_state": quantity_state,
        "assembly_quantity": None,
        "assembly_quantity_state": "blank",
        "effective_assembly_quantity": None,
        "effective_assembly_quantity_state": "blank",
    }


def project_data(rows: list[dict]) -> dict:
    return {
        "project": {"name": "Nested"},
        "subtypes": SUBTYPES,
        "categories": [
            {
                "name": "Category",
                "depth": 0,
                "instances": [
                    {
                        "id": 20,
                        "name": "Instance",
                        "short_name": None,
                        "materials": [
                            {
                                "material_id": 30,
                                "material_name": "Material",
                                "sku": "MAT-1",
                                "unit": "UN",
                                "source_status": "catalog",
                                "bom_entries": rows,
                            }
                        ],
                    }
                ],
            }
        ],
    }


class SubtypeCompactionTests(unittest.TestCase):
    def test_helper_collapses_all_nested_variants_through_a_group(self) -> None:
        rows = [bom_row(1, 5), bom_row(2, 5), bom_row(3, 5)]

        compacted = compact_subtype_rows(rows, SUBTYPES, signature=lambda row: (row["quantity_state"], row["quantity"]))

        self.assertEqual([row["subtype_id"] for row in compacted], [1])

    def test_helper_compacts_rendered_paths_without_subtype_metadata(self) -> None:
        rows = [
            {**bom_row(1, 5), "subtype_id": None},
            {**bom_row(2, 5), "subtype_id": None},
            {**bom_row(3, 5), "subtype_id": None},
        ]

        compacted = compact_subtype_rows(rows, [], signature=lambda row: (row["quantity_state"], row["quantity"]))

        self.assertEqual([row["subtype"] for row in compacted], ["Base"])

    def test_helper_keeps_difference_and_recursively_collapses_its_uniform_children(self) -> None:
        rows = [bom_row(1, 5), bom_row(2, 7), bom_row(3, 7)]

        compacted = compact_subtype_rows(rows, SUBTYPES, signature=lambda row: (row["quantity_state"], row["quantity"]))

        self.assertEqual([row["subtype_id"] for row in compacted], [1, 2])

    def test_helper_keeps_blank_and_zero_distinct(self) -> None:
        rows = [bom_row(1, None), bom_row(2, 0), bom_row(3, 0)]

        compacted = compact_subtype_rows(rows, SUBTYPES, signature=lambda row: (row["quantity_state"], row["quantity"]))

        self.assertEqual([row["subtype_id"] for row in compacted], [1, 2])

    def test_detailed_material_pdf_projection_compacts_aggregated_rows(self) -> None:
        data = project_data([bom_row(1, 5), bom_row(2, 5), bom_row(3, 5)])

        sections = build_detailed_material_export_sections(data)

        self.assertEqual(sections[0]["materials"][0]["rows"], [{"subtype_id": 1, "subtype": "Base", "quantity": 5.0}])

    def test_pdf_projections_compact_rendered_paths_without_subtype_metadata(self) -> None:
        data = project_data(
            [
                {**bom_row(1, 5), "subtype_id": None},
                {**bom_row(2, 5), "subtype_id": None},
                {**bom_row(3, 5), "subtype_id": None},
            ]
        )
        data["subtypes"] = []

        detailed_sections = build_detailed_material_export_sections(data)
        self.assertEqual(
            detailed_sections[0]["materials"][0]["rows"],
            [{"subtype_id": None, "subtype": "Base", "quantity": 5.0}],
        )

        instance = data["categories"][0]["instances"][0]
        technical_materials = _technical_materials(instance, {"include_materials": True}, [])
        self.assertEqual(
            technical_materials[0]["rows"],
            [{"subtype": "Base", "quantity": 5}],
        )

    def test_technical_pdf_projection_keeps_parent_and_distinct_child(self) -> None:
        data = project_data([bom_row(1, 5), bom_row(2, 7), bom_row(3, 7)])
        instance = data["categories"][0]["instances"][0]

        materials = _technical_materials(instance, {"include_materials": True}, SUBTYPES)

        self.assertEqual(
            materials[0]["rows"],
            [
                {"subtype": "Base", "quantity": 5},
                {"subtype": "Base › Organization › Child", "quantity": 7},
            ],
        )


    def test_materials_workbook_compacts_each_quantity_basis_independently(self) -> None:
        rows = []
        for subtype_id, assembly_quantity in ((1, 1), (2, 2), (3, 2)):
            row = bom_row(subtype_id, 1)
            row.update(
                {
                    "assembly_quantity": assembly_quantity,
                    "assembly_quantity_state": "value",
                    "effective_assembly_quantity": assembly_quantity,
                    "effective_assembly_quantity_state": "value",
                }
            )
            rows.append(row)

        output = BytesIO()
        build_materials_workbook(project_data(rows), output)
        output.seek(0)
        workbook = load_workbook(output)

        totals = workbook["Total Materiales"]
        self.assertEqual(
            [totals.cell(row=2, column=column).value for column in range(1, 6)],
            ["Material", "MAT-1", "Base", 1, "UN"],
        )
        self.assertEqual(totals.max_row, 2)

        factory = workbook["Por Contexto"]
        factory_rows = [
            [factory.cell(row=row_index, column=column).value for column in range(1, 6)]
            for row_index in range(3, factory.max_row + 1)
            if isinstance(factory.cell(row=row_index, column=4).value, (int, float))
        ]
        self.assertEqual(factory_rows, [["Material", "MAT-1", "Base", 1, "UN"]])

        work = workbook["Q obra"]
        work_row_indexes = [
            row_index
            for row_index in range(3, work.max_row + 1)
            if isinstance(work.cell(row=row_index, column=4).value, (int, float))
        ]
        work_rows = [
            [work.cell(row=row_index, column=column).value for column in range(1, 6)]
            for row_index in work_row_indexes
        ]
        self.assertEqual(
            work_rows,
            [
                ["Material", "MAT-1", "Base", 1, "UN"],
                [None, None, "Base › Organization › Child", 2, None],
            ],
        )
        self.assertEqual(work.row_dimensions[work_row_indexes[1]].outlineLevel, 2)
        self.assertEqual(work.cell(row=work_row_indexes[0], column=1).border.top.style, "thin")

    def test_materials_workbook_aggregates_totals_before_compacting(self) -> None:
        data = project_data([bom_row(1, 1), bom_row(2, 1), bom_row(3, 1)])
        first_instance = data["categories"][0]["instances"][0]
        data["categories"][0]["instances"].append(
            {
                **first_instance,
                "id": 21,
                "name": "Second Instance",
                "materials": [
                    {
                        **first_instance["materials"][0],
                        "bom_entries": [bom_row(1, 2), bom_row(2, 3), bom_row(3, 3)],
                    }
                ],
            }
        )

        output = BytesIO()
        build_materials_workbook(data, output)
        output.seek(0)
        totals = load_workbook(output)["Total Materiales"]

        self.assertEqual(
            [
                [totals.cell(row=row_index, column=column).value for column in range(1, 6)]
                for row_index in range(2, totals.max_row + 1)
            ],
            [
                ["Material", "MAT-1", "Base", 3, "UN"],
                [None, None, "Base › Organization › Child", 4, None],
            ],
        )

    def test_materials_workbook_omits_blank_parents_when_children_have_values(self) -> None:
        data = project_data([bom_row(1, None), bom_row(2, 1), bom_row(3, 1)])

        output = BytesIO()
        build_materials_workbook(data, output)
        output.seek(0)
        workbook = load_workbook(output)

        totals = workbook["Total Materiales"]
        self.assertEqual(
            [
                [totals.cell(row=row_index, column=column).value for column in range(1, 6)]
                for row_index in range(2, totals.max_row + 1)
            ],
            [["Material", "MAT-1", "Base › Organization › Child", 1, "UN"]],
        )

        factory = workbook["Por Contexto"]
        factory_subtypes = [
            factory.cell(row=row_index, column=3).value
            for row_index in range(3, factory.max_row + 1)
            if factory.cell(row=row_index, column=3).value
        ]
        self.assertEqual(factory_subtypes, ["Base › Organization › Child"])


if __name__ == "__main__":
    unittest.main()
