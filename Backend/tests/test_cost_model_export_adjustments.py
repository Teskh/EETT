from __future__ import annotations

from io import BytesIO
import unittest

from openpyxl import load_workbook

from app.services.export_workbooks import _build_cost_model_rows, build_cost_model_workbook


class CostModelExportAdjustmentTests(unittest.TestCase):
    def setUp(self) -> None:
        self.project_data = {
            "project": {"name": "Proyecto de prueba"},
            "subtypes": [{"id": 9, "name": "Premium", "children": []}],
            "categories": [
                {
                    "number": "1",
                    "name": "Estructura",
                    "instances": [
                        {
                            "id": 10,
                            "name": "Muro",
                            "short_name": "M-01",
                            "materials": [
                                {
                                    "material_id": 21,
                                    "material_name": "Tablero",
                                    "sku": "MAT-021",
                                    "unit": "un",
                                    "bom_entries": [
                                        {
                                            "subtype": "General",
                                            "subtype_id": None,
                                            "quantity": 2,
                                            "quantity_state": "value",
                                        },
                                        {
                                            "subtype": "Premium",
                                            "subtype_id": 9,
                                            "quantity": 3,
                                            "quantity_state": "value",
                                        },
                                    ],
                                }
                            ],
                        }
                    ],
                }
            ],
            "auxiliary_materials": [],
        }
        self.adjustments = [
            {"material_id": 21, "subtype_id": None, "adjusted_quantity": 5},
            {"material_id": 21, "subtype_id": 9, "adjusted_quantity": 1},
        ]

    def test_adjustments_are_exported_as_explicit_deltas(self) -> None:
        rows = _build_cost_model_rows(
            self.project_data,
            prices_by_sku={"MAT-021": 1000},
            adjustments=self.adjustments,
        )

        adjustment_rows = [row for row in rows if row["is_adjustment"]]
        self.assertEqual(len(adjustment_rows), 2)
        deltas = {row["subtype_id"]: row["quantity"] for row in adjustment_rows}
        self.assertEqual(deltas, {None: 3, 9: -2})
        self.assertTrue(all(row["instance_name"] == "Ajustes del modelo" for row in adjustment_rows))

    def test_workbook_contains_adjustment_rows_and_preserves_zero_totals(self) -> None:
        zero_adjustment = [{"material_id": 21, "subtype_id": None, "adjusted_quantity": 0}]
        output = BytesIO()
        build_cost_model_workbook(
            self.project_data,
            output,
            prices_by_sku={"MAT-021": 1000},
            adjustments=zero_adjustment,
        )
        output.seek(0)

        workbook = load_workbook(output, data_only=False)
        by_instance = workbook["Por Instancia"]
        adjustment_row = next(
            row
            for row in by_instance.iter_rows(min_row=2, values_only=True)
            if row[0] == "Ajustes del modelo" and row[2] == "General"
        )
        self.assertEqual(adjustment_row[6], -2)

        totals = workbook["Total Materiales"]
        general_row = next(
            row_index
            for row_index in range(2, totals.max_row + 1)
            if totals.cell(row=row_index, column=1).value == "MAT-021"
            and totals.cell(row=row_index, column=3).value == "General"
        )
        quantity_formula = totals.cell(row=general_row, column=5).value
        self.assertTrue(quantity_formula.startswith("=IF(COUNTIFS"))
        self.assertIn("SUMIFS('Por Instancia'!$G:$G", quantity_formula)


if __name__ == "__main__":
    unittest.main()
