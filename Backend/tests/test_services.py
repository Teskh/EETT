from __future__ import annotations

import asyncio
from datetime import date, datetime, timedelta, timezone
import importlib.util
from io import BytesIO
import os
import tempfile
from types import SimpleNamespace
import unittest
from unittest.mock import patch

from openpyxl import load_workbook
from fastapi.testclient import TestClient
from sqlalchemy import select
from sqlalchemy.exc import OperationalError
from sqlalchemy.orm import sessionmaker

from app.config import Settings
from app.database import Base, create_engine_for_url
from app.main import create_app
from app.models import (
    CatalogCategory,
    CatalogComponent,
    ComponentType,
    ComponentMaterialRule,
    ErpMaterialCache,
    MaterialDashboardCacheEntry,
    MaterialStudyGroup,
    Project,
    ProjectActivityGroup,
    ProjectActivityLog,
    ProjectBomEntry,
    ProjectInstance,
    ProjectSubtype,
)
from app.seed import seed_demo_data_if_empty
from app.services.catalog import (
    create_attribute_definition,
    create_category,
    create_component,
    delete_attribute_definition,
    get_catalog_page_data,
    replace_component_material_rules,
    replace_component_attributes,
    search_material_candidates,
    update_attribute_definition,
)
from app.services.audit import build_activity_details
from app.services.dashboard import get_material_dashboard_economic_metrics, get_material_dashboard_history, get_recent_material_dashboard
from app.services.dashboard import _add_business_days, _build_material_dashboard_detail, _count_business_days
from app.services.erp import (
    _calculate_delivery_time_stats,
    _get_last_purchase_orders_for_products_batch,
    _get_lead_time_samples_for_product,
)
from app.services.material_groups import (
    create_material_study_group,
    get_material_dashboard_group_detail,
    get_material_dashboard_group_history,
    get_material_dashboard_groups,
    update_material_study_group,
)
from app.services.projects import (
    _visible_project_subtype_rows,
    create_project_instance,
    get_instance_sync_preview,
    get_project_view_data,
    get_projects_page_data,
    refresh_instance_snapshot,
)
from app.ui import render_catalog_page, render_project_detail_page, render_projects_page


class ServiceLayerTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.test_database_url = os.getenv("SPEC_SHEETS_TEST_DATABASE_URL")
        if not cls.test_database_url:
            raise unittest.SkipTest("SPEC_SHEETS_TEST_DATABASE_URL is not set")

    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.settings = Settings(
            database_url=self.test_database_url,
            seed_demo_data=True,
            environment="test",
            allow_trusted_user_header=True,
            export_output_dir=self.temp_dir.name,
        )
        self.engine = create_engine_for_url(self.settings.database_url)
        self.session_factory = sessionmaker(bind=self.engine, autoflush=False, expire_on_commit=False)
        Base.metadata.drop_all(self.engine)
        Base.metadata.create_all(self.engine)
        seed_demo_data_if_empty(self.session_factory)
        self.app = create_app(
            Settings(
                database_url=self.test_database_url,
                seed_demo_data=False,
                require_schema=True,
                environment="test",
                allow_trusted_user_header=True,
                export_output_dir=self.temp_dir.name,
            )
        )
        self.client = TestClient(self.app)

    def tearDown(self) -> None:
        self.client.close()
        self.engine.dispose()
        self.temp_dir.cleanup()

    def test_catalog_page_data_contains_expected_demo_structure(self) -> None:
        with self.session_factory() as session:
            data = get_catalog_page_data(session)

        self.assertEqual(data["summary"]["categories"], 14)
        self.assertEqual(data["selected"]["name"], "Openings")
        child_names = {child["name"] for child in data["selected"]["child_categories"]}
        self.assertIn("Doors", child_names)
        self.assertIn("Windows", child_names)

    def test_catalog_create_helpers_persist_new_records(self) -> None:
        with self.session_factory() as session:
            category = create_category(
                session,
                name="Bathrooms",
                description="Wet-area catalog lines",
                scope="item",
                parent_id=None,
            )
            component = create_component(
                session,
                category_id=category.id,
                component_type="item",
                name="Mirror Cabinet",
                short_name="MC-01",
                description="Wall-mounted cabinet",
                short_description="Client-facing mirror cabinet",
                installation="Anchor to masonry wall.",
                unit_type="unit",
            )
            refreshed = get_catalog_page_data(session, selected_category_id=category.id)

        self.assertEqual(component.name, "Mirror Cabinet")
        self.assertEqual(component.short_description, "Client-facing mirror cabinet")
        self.assertEqual(refreshed["selected"]["name"], "Bathrooms")
        self.assertEqual(len(refreshed["selected"]["components"]), 1)
        self.assertEqual(refreshed["selected"]["components"][0]["short_description"], "Client-facing mirror cabinet")

    def test_catalog_attribute_helpers_support_create_update_delete_and_bump_sync_timestamp(self) -> None:
        with self.session_factory() as session:
            component = session.scalar(select(CatalogComponent).where(CatalogComponent.name == "Base Cabinet 900"))
            project = session.scalar(select(Project).where(Project.name == "Casa Robles - Block A"))

            self.assertIsNotNone(component)
            self.assertIsNotNone(project)
            assert component is not None
            assert project is not None

            instance = create_project_instance(
                session,
                project=project,
                category_id=component.category_id,
                component_id=component.id,
                name="Kitchen Cabinet Clone",
                short_name="KCC-01",
                description="Snapshot for sync test",
                short_description="Short sync test copy",
                installation="Install per kitchen plan.",
                unit_amount=1,
            )
            baseline_sync = get_instance_sync_preview(session, instance.id)
            self.assertIsNotNone(baseline_sync)
            self.assertFalse(baseline_sync["is_outdated"])

            definition = create_attribute_definition(
                session,
                component_id=component.id,
                name="Toe Kick Finish",
                value_type="select",
                options_text="Oak, Walnut\nGraphite",
            )
            self.assertIsNotNone(definition)
            assert definition is not None
            self.assertEqual([option.value for option in definition.options], ["Oak", "Walnut", "Graphite"])

            refreshed_sync = get_instance_sync_preview(session, instance.id)
            self.assertIsNotNone(refreshed_sync)
            self.assertTrue(refreshed_sync["is_outdated"])

            updated = update_attribute_definition(
                session,
                attribute_definition_id=definition.id,
                name="Toe Kick Material",
                value_type="text",
                options_text="Ignored",
            )
            self.assertIsNotNone(updated)
            assert updated is not None
            self.assertEqual(updated.name, "Toe Kick Material")
            self.assertEqual(updated.value_type.value, "text")
            self.assertEqual(updated.options, [])

            deleted_category_id = delete_attribute_definition(session, attribute_definition_id=definition.id)
            self.assertEqual(deleted_category_id, component.category_id)

            refreshed = get_catalog_page_data(session, selected_category_id=component.category_id)
            component_payload = next(item for item in refreshed["selected"]["components"] if item["id"] == component.id)
            self.assertNotIn("Toe Kick Material", [attribute["name"] for attribute in component_payload["attributes"]])

    def test_replace_component_attributes_preserves_nested_option_rows_and_marks_instances_outdated(self) -> None:
        with self.session_factory() as session:
            component = session.scalar(select(CatalogComponent).where(CatalogComponent.name == "Base Cabinet 900"))
            project = session.scalar(select(Project).where(Project.name == "Casa Robles - Block A"))

            self.assertIsNotNone(component)
            self.assertIsNotNone(project)
            assert component is not None
            assert project is not None

            instance = create_project_instance(
                session,
                project=project,
                category_id=component.category_id,
                component_id=component.id,
                name="Kitchen Cabinet UX",
                short_name="KCU-01",
                description="Snapshot for bulk attribute editor test",
                short_description="Bulk editor snapshot",
                installation="Install per kitchen plan.",
                unit_amount=1,
            )
            self.assertFalse(get_instance_sync_preview(session, instance.id)["is_outdated"])

            updated_component = replace_component_attributes(
                session,
                component_id=component.id,
                attributes=[
                    {
                        "name": "Countertop",
                        "value_type": "select",
                        "options": ["Laminate", "Quartz", "Porcelain"],
                    },
                    {
                        "name": "Depth",
                        "value_type": "number",
                        "options": ["ignored"],
                    },
                ],
            )

            self.assertIsNotNone(updated_component)
            refreshed = get_catalog_page_data(session, selected_category_id=component.category_id)
            component_payload = next(item for item in refreshed["selected"]["components"] if item["id"] == component.id)
            self.assertEqual(
                component_payload["attributes"],
                [
                    {
                        "id": component_payload["attributes"][0]["id"],
                        "name": "Countertop",
                        "value_type": "select",
                        "options": ["Laminate", "Quartz", "Porcelain"],
                    },
                    {
                        "id": component_payload["attributes"][1]["id"],
                        "name": "Depth",
                        "value_type": "number",
                        "options": [],
                    },
                ],
            )
            self.assertTrue(get_instance_sync_preview(session, instance.id)["is_outdated"])

    def test_replace_component_material_rules_marks_removed_materials_stale_without_deleting_project_bom(self) -> None:
        with self.session_factory() as session:
            component = session.scalar(select(CatalogComponent).where(CatalogComponent.name == "Entry Door"))
            project = session.scalar(select(Project).where(Project.name == "Casa Robles - Block A"))
            door_instance = next(
                instance for instance in project.instances if instance.name == "Door A"
            ) if project is not None else None

            self.assertIsNotNone(component)
            self.assertIsNotNone(project)
            self.assertIsNotNone(door_instance)
            assert component is not None
            assert project is not None
            assert door_instance is not None
            retained_rule = next(rule for rule in component.material_rules if rule.material.sku == "MAT-001")

            updated_component = replace_component_material_rules(
                session,
                component_id=component.id,
                rules=[
                    {
                        "id": retained_rule.id,
                        "material_name": "Anchor Screw 5x70",
                        "sku": "MAT-001",
                        "unit": "ea",
                        "unit_qty_per_unit": 8,
                        "notes": "Door still needs fixing hardware.",
                        "conditions": [],
                    }
                ],
            )

            self.assertIsNotNone(updated_component)

            remaining_bom_rows = session.scalars(
                select(ProjectBomEntry)
                .where(ProjectBomEntry.instance_id == door_instance.id)
                .order_by(ProjectBomEntry.material_id)
            ).all()
            self.assertEqual([row.material.sku for row in remaining_bom_rows], ["MAT-001", "MAT-002"])

            detail = get_project_view_data(session, project.id)
            self.assertIsNotNone(detail)
            assert detail is not None
            door_section = next(section for section in detail["categories"] if section["name"] == "Doors")
            door_instance = next(item for item in door_section["instances"] if item["name"] == "Door A")
            self.assertEqual([material["sku"] for material in door_instance["materials"]], ["MAT-001", "MAT-002"])
            self.assertEqual(next(material for material in door_instance["materials"] if material["sku"] == "MAT-002")["source_status"], "stale")

            catalog_payload = get_catalog_page_data(session, selected_category_id=component.category_id)
            selected_component = next(
                item for item in catalog_payload["selected"]["components"] if item["id"] == component.id
            )
            self.assertIsNotNone(selected_component["material_rules"][0]["id"])
            self.assertEqual(selected_component["material_rules"][0]["id"], retained_rule.id)

    def test_visible_project_subtype_rows_ignore_flat_subtypes_outside_visible_tree(self) -> None:
        project = Project(name="Subtype Visibility Contract", status="template")
        root = ProjectSubtype(project=project, name="Root")
        root.id = 1
        visible_child = ProjectSubtype(project=project, name="Visible Child", parent=root)
        visible_child.id = 2
        visible_child.parent_id = root.id
        dangling = ProjectSubtype(project=project, name="Dangling", parent_id=999999)
        dangling.id = 3

        rows = _visible_project_subtype_rows(project)

        self.assertEqual(
            [(row["name"], row["depth"]) for row in rows],
            [("Root", 0), ("Visible Child", 1)],
        )

    def test_catalog_material_api_supports_search_and_bulk_rule_save(self) -> None:
        with self.session_factory() as session:
            component = session.scalar(select(CatalogComponent).where(CatalogComponent.name == "Sliding Window"))
            self.assertIsNotNone(component)
            assert component is not None
            component_id = component.id
            category_id = component.category_id

            local_results = search_material_candidates(session, query="MAT-00")
            self.assertTrue(any(result["sku"] == "MAT-003" for result in local_results))

        search_response = self.client.get(
            "/api/v1/catalog/materials/search?q=MAT-003",
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(search_response.status_code, 200)
        self.assertTrue(any(result["sku"] == "MAT-003" for result in search_response.json()["results"]))

        replace_response = self.client.put(
            f"/api/v1/catalog/components/{component_id}/materials",
            headers={"X-Spec-Sheets-User": "editor"},
            json={
                "rules": [
                    {
                        "material_name": "Neutral Cure Silicone",
                        "sku": "MAT-006",
                        "unit": "cartridge",
                        "unit_qty_per_unit": 0.25,
                        "notes": "Seal perimeter after glazing adjustment.",
                        "conditions": [],
                    }
                ]
            },
        )
        self.assertEqual(replace_response.status_code, 200)
        self.assertEqual(replace_response.json()["component_id"], component_id)

        catalog_response = self.client.get(
            f"/api/v1/catalog?category_id={category_id}",
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(catalog_response.status_code, 200)
        selected_component = next(
            item for item in catalog_response.json()["selected"]["components"] if item["id"] == component_id
        )
        self.assertEqual([rule["sku"] for rule in selected_component["material_rules"]], ["MAT-006"])
        self.assertIsNotNone(selected_component["material_rules"][0]["id"])

    def test_catalog_and_project_crud_routes_work(self) -> None:
        create_catalog_component_response = self.client.post(
            "/catalog/components",
            headers={"X-Spec-Sheets-User": "editor"},
            data={
                "category_id": "6",
                "component_type": "item",
                "name": "Island Module",
                "short_name": "ISL-01",
                "description": "Kitchen island module",
                "installation": "Set and level on finished floor.",
                "unit_type": "unit",
            },
            follow_redirects=False,
        )
        self.assertEqual(create_catalog_component_response.status_code, 303)

        with self.session_factory() as session:
            kitchens = get_catalog_page_data(session, selected_category_id=6)
            component = next(item for item in kitchens["selected"]["components"] if item["name"] == "Island Module")
            component_id = component["id"]

        update_catalog_component_response = self.client.post(
            f"/catalog/components/{component_id}/update",
            headers={"X-Spec-Sheets-User": "editor"},
            data={
                "name": "Island Module Revised",
                "short_name": "ISL-02",
                "description": "Updated module",
                "installation": "Install after plumbing rough-in.",
                "unit_type": "set",
                "component_type": "item",
            },
            follow_redirects=False,
        )
        self.assertEqual(update_catalog_component_response.status_code, 303)

        create_instance_response = self.client.post(
            "/projects/2/instances",
            headers={"X-Spec-Sheets-User": "editor"},
            data={
                "category_id": "6",
                "component_id": str(component_id),
                "name": "Kitchen Island A",
                "short_name": "KIA-01",
                "description": "Execution instance",
                "installation": "Install on site.",
                "unit_amount": "2",
                "attribute_values_json": '[{"name":"Countertop","value":"Quartz"}]',
            },
            follow_redirects=False,
        )
        self.assertEqual(create_instance_response.status_code, 303)

        project_detail = self.client.get("/api/v1/projects/2", headers={"X-Spec-Sheets-User": "editor"}).json()
        kitchen_section = next(section for section in project_detail["categories"] if section["name"] == "Kitchens")
        created_instance = next(item for item in kitchen_section["instances"] if item["name"] == "Kitchen Island A")

        update_instance_response = self.client.post(
            f"/projects/2/instances/{created_instance['id']}/update",
            headers={"X-Spec-Sheets-User": "editor"},
            data={
                "category_id": "6",
                "name": "Kitchen Island B",
                "short_name": "KIB-01",
                "description": "Updated execution instance",
                "installation": "Updated install notes.",
                "unit_amount": "3",
                "attribute_values_json": '[{"name":"Countertop","value":"Laminate"}]',
            },
            follow_redirects=False,
        )
        self.assertEqual(update_instance_response.status_code, 303)

        project_detail_after_update = self.client.get("/api/v1/projects/2", headers={"X-Spec-Sheets-User": "editor"}).json()
        kitchen_section_after_update = next(section for section in project_detail_after_update["categories"] if section["name"] == "Kitchens")
        updated_instance = next(item for item in kitchen_section_after_update["instances"] if item["name"] == "Kitchen Island B")
        self.assertEqual(updated_instance["attributes"][0]["values"][0]["value"], "Laminate")

        delete_instance_response = self.client.post(
            f"/projects/2/instances/{created_instance['id']}/delete",
            headers={"X-Spec-Sheets-User": "editor"},
            data={"category_id": "6"},
            follow_redirects=False,
        )
        self.assertEqual(delete_instance_response.status_code, 303)

    def test_catalog_attribute_routes_work(self) -> None:
        with self.session_factory() as session:
            component = session.scalar(select(CatalogComponent).where(CatalogComponent.name == "Base Cabinet 900"))
            self.assertIsNotNone(component)
            assert component is not None
            component_id = component.id
            category_id = component.category_id

        create_response = self.client.post(
            f"/catalog/components/{component_id}/attributes",
            headers={"X-Spec-Sheets-User": "editor"},
            data={
                "name": "Drawer Front",
                "value_type": "select",
                "options_text": "Slab, Shaker",
            },
            follow_redirects=False,
        )
        self.assertEqual(create_response.status_code, 303)

        with self.session_factory() as session:
            refreshed = get_catalog_page_data(session, selected_category_id=category_id)
            component_payload = next(item for item in refreshed["selected"]["components"] if item["id"] == component_id)
            attribute = next(item for item in component_payload["attributes"] if item["name"] == "Drawer Front")

        update_response = self.client.post(
            f"/catalog/attributes/{attribute['id']}/update",
            headers={"X-Spec-Sheets-User": "editor"},
            data={
                "name": "Drawer Style",
                "value_type": "text",
                "options_text": "Should be cleared",
            },
            follow_redirects=False,
        )
        self.assertEqual(update_response.status_code, 303)

        delete_response = self.client.post(
            f"/catalog/attributes/{attribute['id']}/delete",
            headers={"X-Spec-Sheets-User": "editor"},
            data={"category_id": str(category_id)},
            follow_redirects=False,
        )
        self.assertEqual(delete_response.status_code, 303)

        with self.session_factory() as session:
            refreshed = get_catalog_page_data(session, selected_category_id=category_id)
            component_payload = next(item for item in refreshed["selected"]["components"] if item["id"] == component_id)
            self.assertNotIn("Drawer Style", [item["name"] for item in component_payload["attributes"]])

    def test_catalog_bulk_attribute_route_matches_legacy_repeatable_editor_flow(self) -> None:
        with self.session_factory() as session:
            component = session.scalar(select(CatalogComponent).where(CatalogComponent.name == "Base Cabinet 900"))
            self.assertIsNotNone(component)
            assert component is not None
            component_id = component.id
            category_id = component.category_id

        response = self.client.post(
            f"/catalog/components/{component_id}/attributes/update",
            headers={"X-Spec-Sheets-User": "editor"},
            data={
                "attributes_json": '[{"name":"Drawer Style","value_type":"select","options":["Slab","Shaker"]},{"name":"Cabinet Width","value_type":"number","options":["should clear"]}]',
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)

        with self.session_factory() as session:
            refreshed = get_catalog_page_data(session, selected_category_id=category_id)
            component_payload = next(item for item in refreshed["selected"]["components"] if item["id"] == component_id)
            self.assertEqual(
                [(item["name"], item["value_type"], item["options"]) for item in component_payload["attributes"]],
                [
                    ("Drawer Style", "select", ["Slab", "Shaker"]),
                    ("Cabinet Width", "number", []),
                ],
            )

    def test_catalog_bulk_attribute_route_supports_fetch_without_redirect(self) -> None:
        with self.session_factory() as session:
            component = session.scalar(select(CatalogComponent).where(CatalogComponent.name == "Base Cabinet 900"))
            self.assertIsNotNone(component)
            assert component is not None

        response = self.client.post(
            f"/catalog/components/{component.id}/attributes/update",
            headers={
                "X-Spec-Sheets-User": "editor",
                "x-requested-with": "fetch",
            },
            data={
                "attributes_json": '[{"name":"Drawer Style","value_type":"select","options":["Slab","Shaker"]}]',
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["ok"], True)

    def test_project_view_marks_instance_outdated_after_catalog_attribute_change_and_refresh_adds_missing_rows(self) -> None:
        with self.session_factory() as session:
            component = session.scalar(select(CatalogComponent).where(CatalogComponent.name == "Base Cabinet 900"))
            project = session.scalar(select(Project).where(Project.name == "Casa Robles - Block A"))

            self.assertIsNotNone(component)
            self.assertIsNotNone(project)
            assert component is not None
            assert project is not None

            instance = create_project_instance(
                session,
                project=project,
                category_id=component.category_id,
                component_id=component.id,
                name="Kitchen Cabinet Snapshot",
                short_name="KCS-01",
                description="Snapshot to test attribute refresh",
                short_description="Refresh snapshot",
                installation="Install per kitchen plan.",
                unit_amount=1,
            )
            baseline = get_project_view_data(session, project.id)
            kitchens = next(section for section in baseline["categories"] if section["name"] == "Kitchens")
            baseline_instance = next(item for item in kitchens["instances"] if item["id"] == instance.id)
            self.assertFalse(baseline_instance["sync_state"]["is_outdated"])
            self.assertEqual(baseline_instance["short_description"], "Refresh snapshot")
            self.assertEqual(
                [row["name"] for row in baseline_instance["attributes"][0]["values"]],
                ["Countertop"],
            )

            replace_component_attributes(
                session,
                component_id=component.id,
                attributes=[
                    {
                        "name": "Countertop",
                        "value_type": "select",
                        "options": ["Laminate", "Quartz"],
                    },
                    {
                        "name": "Depth",
                        "value_type": "number",
                        "options": [],
                    },
                ],
            )

            changed = get_project_view_data(session, project.id)
            kitchens = next(section for section in changed["categories"] if section["name"] == "Kitchens")
            changed_instance = next(item for item in kitchens["instances"] if item["id"] == instance.id)
            self.assertTrue(changed_instance["sync_state"]["is_outdated"])
            self.assertEqual(changed_instance["sync_state"]["status"], "out_of_sync")
            self.assertEqual(
                [row["name"] for row in changed_instance["attributes"][0]["values"]],
                ["Countertop"],
            )

            refreshed_preview = refresh_instance_snapshot(session, instance_id=instance.id, actor_user=None)
            self.assertIsNotNone(refreshed_preview)
            self.assertFalse(refreshed_preview["is_outdated"])

            refreshed = get_project_view_data(session, project.id)
            kitchens = next(section for section in refreshed["categories"] if section["name"] == "Kitchens")
            refreshed_instance = next(item for item in kitchens["instances"] if item["id"] == instance.id)
            self.assertFalse(refreshed_instance["sync_state"]["is_outdated"])
            self.assertEqual(
                [row["name"] for row in refreshed_instance["attributes"][0]["values"]],
                ["Countertop", "Depth"],
            )

    def test_projects_page_and_detail_preserve_material_semantics(self) -> None:
        with self.session_factory() as session:
            board = get_projects_page_data(session)
            execution_project = board["grouped_projects"]["execution"][0]
            detail = get_project_view_data(session, execution_project["id"])

        self.assertEqual(execution_project["name"], "Casa Robles - Block A")
        self.assertEqual(execution_project["material_mode"], "per_subtype")
        self.assertEqual(detail["project"]["status"], "execution")

        window_section = next(section for section in detail["categories"] if section["name"] == "Windows")
        window_instance = window_section["instances"][0]
        silicone_material = next(material for material in window_instance["materials"] if material["sku"] == "MAT-006")
        self.assertEqual(silicone_material["bom_entries"][0]["quantity_state"], "blank")

        trim_section = next(section for section in detail["categories"] if section["name"] == "Trim")
        trim_instance = trim_section["instances"][0]
        trim_material = trim_instance["materials"][0]
        self.assertEqual(trim_material["bom_entries"][0]["quantity_state"], "zero")
        self.assertEqual(trim_instance["sync_state"]["status"], "out_of_sync")

    def test_project_view_includes_explicit_usage_occurrences_for_linked_accessories(self) -> None:
        with self.session_factory() as session:
            detail = get_project_view_data(session, 2)

        self.assertIsNotNone(detail)
        assert detail is not None

        paints_section = next(section for section in detail["categories"] if section["name"] == "Paints")
        paint_instance = next(item for item in paints_section["instances"] if item["name"] == "Exterior Enamel Paint")
        self.assertEqual(len(paint_instance["outgoing_occurrences"]), 3)
        self.assertEqual(
            [row["context_label"] for row in paint_instance["outgoing_occurrences"]],
            ["Primary railing bars", "Porch timber fascia", "Exterior doorknob set"],
        )

        sealants_section = next(section for section in detail["categories"] if section["name"] == "Sealants")
        caulking_instance = next(item for item in sealants_section["instances"] if item["name"] == "Elastic Caulking Package")
        freeform_occurrence = next(
            row for row in caulking_instance["outgoing_occurrences"] if row["context_label"] == "Joint between kitchen wall and ceiling"
        )
        self.assertEqual(freeform_occurrence["targets"], [])

        railings_section = next(section for section in detail["categories"] if section["name"] == "Railings")
        railing_instance = next(item for item in railings_section["instances"] if item["name"] == "Main Railing")
        self.assertEqual(len(railing_instance["incoming_occurrences"]), 1)
        self.assertEqual(railing_instance["incoming_occurrences"][0]["context_label"], "Primary railing bars")

    def test_project_view_exposes_usage_attribute_definitions_for_occurrence_editing(self) -> None:
        with self.session_factory() as session:
            detail = get_project_view_data(session, 2)

        self.assertIsNotNone(detail)
        assert detail is not None

        sealants_section = next(section for section in detail["categories"] if section["name"] == "Sealants")
        caulking_instance = next(item for item in sealants_section["instances"] if item["name"] == "Elastic Caulking Package")
        self.assertEqual(
            [row["name"] for row in caulking_instance["usage_attribute_definitions"]],
            ["Color", "Applicability", "Area"],
        )

    def test_project_occurrence_api_supports_create_update_and_delete(self) -> None:
        with self.session_factory() as session:
            caulking_instance = session.scalar(
                select(ProjectInstance).where(ProjectInstance.project_id == 2, ProjectInstance.name == "Elastic Caulking Package")
            )
            toilet_instance = session.scalar(
                select(ProjectInstance).where(ProjectInstance.project_id == 2, ProjectInstance.name == "Bathroom Toilet")
            )

        self.assertIsNotNone(caulking_instance)
        self.assertIsNotNone(toilet_instance)
        assert caulking_instance is not None
        assert toilet_instance is not None

        create_response = self.client.post(
            f"/api/v1/projects/2/instances/{caulking_instance.id}/occurrences",
            headers={"X-Spec-Sheets-User": "editor"},
            json={
                "relationship_type": "seals",
                "context_label": "Mirror splashback return",
                "target_instance_id": toilet_instance.id,
                "attribute_values": [
                    {"name": "Color", "value": "White"},
                    {"name": "Applicability", "value": "Toilet base to wall"},
                    {"name": "Area", "value": "Mirror return"},
                ],
            },
        )
        self.assertEqual(create_response.status_code, 200)
        occurrence_id = create_response.json()["occurrence_id"]

        detail_after_create = self.client.get("/api/v1/projects/2", headers={"X-Spec-Sheets-User": "editor"})
        self.assertEqual(detail_after_create.status_code, 200)
        sealants_section = next(section for section in detail_after_create.json()["categories"] if section["name"] == "Sealants")
        caulking_payload = next(item for item in sealants_section["instances"] if item["id"] == caulking_instance.id)
        created_occurrence = next(item for item in caulking_payload["outgoing_occurrences"] if item["id"] == occurrence_id)
        self.assertEqual(created_occurrence["targets"][0]["instance_id"], toilet_instance.id)
        self.assertEqual(created_occurrence["attributes"][1]["name"], "Applicability")

        update_response = self.client.put(
            f"/api/v1/projects/2/instances/{caulking_instance.id}/occurrences/{occurrence_id}",
            headers={"X-Spec-Sheets-User": "editor"},
            json={
                "relationship_type": "seals",
                "context_label": "Mirror splashback return revised",
                "target_instance_id": None,
                "attribute_values": [
                    {"name": "Color", "value": "Gray"},
                    {"name": "Applicability", "value": "Wall to ceiling"},
                    {"name": "Area", "value": "Mirror crown line"},
                ],
            },
        )
        self.assertEqual(update_response.status_code, 200)

        detail_after_update = self.client.get("/api/v1/projects/2", headers={"X-Spec-Sheets-User": "editor"})
        self.assertEqual(detail_after_update.status_code, 200)
        sealants_section = next(section for section in detail_after_update.json()["categories"] if section["name"] == "Sealants")
        caulking_payload = next(item for item in sealants_section["instances"] if item["id"] == caulking_instance.id)
        updated_occurrence = next(item for item in caulking_payload["outgoing_occurrences"] if item["id"] == occurrence_id)
        self.assertEqual(updated_occurrence["context_label"], "Mirror splashback return revised")
        self.assertEqual(updated_occurrence["targets"], [])

        delete_response = self.client.delete(
            f"/api/v1/projects/2/instances/{caulking_instance.id}/occurrences/{occurrence_id}",
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(delete_response.status_code, 200)

        detail_after_delete = self.client.get("/api/v1/projects/2", headers={"X-Spec-Sheets-User": "editor"})
        self.assertEqual(detail_after_delete.status_code, 200)
        sealants_section = next(section for section in detail_after_delete.json()["categories"] if section["name"] == "Sealants")
        caulking_payload = next(item for item in sealants_section["instances"] if item["id"] == caulking_instance.id)
        self.assertFalse(any(item["id"] == occurrence_id for item in caulking_payload["outgoing_occurrences"]))

    def test_html_renderers_include_core_screen_content(self) -> None:
        with self.session_factory() as session:
            catalog = get_catalog_page_data(session)
            projects = get_projects_page_data(session)
            project = get_project_view_data(session, 2)

        catalog_html = render_catalog_page(catalog, catalog["selected"]["id"])
        projects_html = render_projects_page(projects)
        project_html = render_project_detail_page(project)

        self.assertIn("Database Editor", catalog_html)
        self.assertIn("Project board", projects_html)
        self.assertIn("Applicable materials", project_html)

    def test_session_and_project_permissions_are_typed_and_filtered(self) -> None:
        session_response = self.client.get("/api/v1/session", headers={"X-Spec-Sheets-User": "viewer"})
        self.assertEqual(session_response.status_code, 200)
        session_payload = session_response.json()
        self.assertEqual(session_payload["username"], "viewer")
        self.assertIn("viewer", session_payload["roles"])
        self.assertFalse(session_payload["permissions"]["catalog_edit"])
        self.assertFalse(session_payload["permissions"]["project_create"])
        self.assertFalse(session_payload["permissions"]["user_admin"])

        projects_response = self.client.get("/api/v1/projects", headers={"X-Spec-Sheets-User": "viewer"})
        self.assertEqual(projects_response.status_code, 200)
        payload = projects_response.json()
        self.assertEqual(payload["grouped_projects"]["template"], [])
        execution_names = [project["name"] for project in payload["grouped_projects"]["execution"]]
        self.assertIn("Casa Robles - Block A", execution_names)

    def test_login_logout_and_user_admin_endpoints_require_sysadmin(self) -> None:
        unauthorized_users = self.client.get("/api/v1/users")
        self.assertEqual(unauthorized_users.status_code, 401)

        bad_login = self.client.post("/api/v1/login", json={"username": "sysadmin", "password": "wrong"})
        self.assertEqual(bad_login.status_code, 401)

        login_response = self.client.post("/api/v1/login", json={"username": "sysadmin", "password": "adminpass"})
        self.assertEqual(login_response.status_code, 200)
        self.assertTrue(login_response.json()["permissions"]["user_admin"])
        self.assertTrue(login_response.json()["permissions"]["project_edit"])

        users_response = self.client.get("/api/v1/users")
        self.assertEqual(users_response.status_code, 200)
        sysadmin_id = next(user["id"] for user in users_response.json()["users"] if user["username"] == "sysadmin")
        roles = {role["code"] for role in users_response.json()["roles"]}
        self.assertEqual(roles, {"admin", "editor", "ot", "viewer"})

        create_user_response = self.client.post(
            "/api/v1/users",
            json={
                "username": "fieldlead",
                "display_name": "Field Lead",
                "email": "fieldlead@specsheets.local",
                "password": "fieldpass",
                "role_codes": ["viewer"],
                "is_active": True,
            },
        )
        self.assertEqual(create_user_response.status_code, 200)
        self.assertEqual(create_user_response.json()["username"], "fieldlead")

        viewer_forbidden = self.client.get("/api/v1/users", headers={"X-Spec-Sheets-User": "viewer"})
        self.assertEqual(viewer_forbidden.status_code, 403)

        update_user_response = self.client.put(
            f"/api/v1/users/{create_user_response.json()['id']}",
            json={
                "display_name": "Field Lead Updated",
                "email": "fieldlead@specsheets.local",
                "password": "fieldpass2",
                "role_codes": ["ot"],
                "is_active": True,
            },
        )
        self.assertEqual(update_user_response.status_code, 200)
        self.assertEqual(update_user_response.json()["roles"], ["ot"])

        logout_response = self.client.post("/api/v1/logout")
        self.assertEqual(logout_response.status_code, 204)

        updated_login = self.client.post("/api/v1/login", json={"username": "fieldlead", "password": "fieldpass2"})
        self.assertEqual(updated_login.status_code, 200)
        self.assertIn("ot", updated_login.json()["roles"])

        delete_with_non_sysadmin = self.client.delete(f"/api/v1/users/{create_user_response.json()['id']}")
        self.assertEqual(delete_with_non_sysadmin.status_code, 403)

        self.client.post("/api/v1/logout")
        relogin_sysadmin = self.client.post("/api/v1/login", json={"username": "sysadmin", "password": "adminpass"})
        self.assertEqual(relogin_sysadmin.status_code, 200)

        delete_user_response = self.client.delete(f"/api/v1/users/{create_user_response.json()['id']}")
        self.assertEqual(delete_user_response.status_code, 200)

        delete_reserved_sysadmin = self.client.delete(f"/api/v1/users/{sysadmin_id}")
        self.assertEqual(delete_reserved_sysadmin.status_code, 422)

    def test_sync_preview_and_material_mode_endpoints(self) -> None:
        mode_response = self.client.get("/api/v1/projects/2/material-mode", headers={"X-Spec-Sheets-User": "editor"})
        self.assertEqual(mode_response.status_code, 200)
        self.assertEqual(mode_response.json()["mode"], "per_subtype")

        preview_response = self.client.get(
            "/api/v1/projects/2/instances/4/sync-preview",
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(preview_response.status_code, 200)
        self.assertTrue(preview_response.json()["is_outdated"])
        self.assertIn("scalar_fields", preview_response.json())
        self.assertIn("attribute_schema", preview_response.json())

        refresh_response = self.client.post(
            "/api/v1/projects/2/instances/4/refresh",
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(refresh_response.status_code, 200)
        self.assertEqual(refresh_response.json()["sync_status"], "up_to_date")
        self.assertFalse(refresh_response.json()["is_outdated"])

        update_mode = self.client.put(
            "/api/v1/projects/2/material-mode",
            headers={"X-Spec-Sheets-User": "editor"},
            json={"mode": "general"},
        )
        self.assertEqual(update_mode.status_code, 200)
        self.assertEqual(update_mode.json()["mode"], "general")

    def test_project_subtype_crud_endpoints(self) -> None:
        create_root = self.client.post(
            "/api/v1/projects/2/subtypes",
            headers={"X-Spec-Sheets-User": "editor"},
            json={"name": "Economy"},
        )
        self.assertEqual(create_root.status_code, 200)
        root_id = create_root.json()["subtype_id"]

        create_child = self.client.post(
            "/api/v1/projects/2/subtypes",
            headers={"X-Spec-Sheets-User": "editor"},
            json={"name": "Economy Child", "parent_id": root_id},
        )
        self.assertEqual(create_child.status_code, 200)

        rename_root = self.client.put(
            f"/api/v1/projects/2/subtypes/{root_id}",
            headers={"X-Spec-Sheets-User": "editor"},
            json={"name": "Economy Revised"},
        )
        self.assertEqual(rename_root.status_code, 200)

        detail = self.client.get("/api/v1/projects/2", headers={"X-Spec-Sheets-User": "editor"})
        self.assertEqual(detail.status_code, 200)
        created_root = next(subtype for subtype in detail.json()["subtypes"] if subtype["id"] == root_id)
        self.assertEqual(created_root["name"], "Economy Revised")
        self.assertEqual(created_root["children"][0]["name"], "Economy Child")

        delete_root = self.client.delete(
            f"/api/v1/projects/2/subtypes/{root_id}",
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(delete_root.status_code, 200)

        with self.session_factory() as session:
            deleted = session.scalar(select(ProjectSubtype).where(ProjectSubtype.id == root_id))
        self.assertIsNone(deleted)

    def test_project_material_occurrence_updates_toggle_between_general_and_per_subtype(self) -> None:
        with self.session_factory() as session:
            window_instance = session.scalar(
                select(ProjectInstance).where(ProjectInstance.project_id == 2, ProjectInstance.name == "Living Window")
            )
            material_rule = session.scalar(
                select(ComponentMaterialRule)
                .where(ComponentMaterialRule.component.has(CatalogComponent.name == "Sliding Window"))
                .where(ComponentMaterialRule.material.has(name="Laminated Glass Panel"))
            )
            subtypes = session.scalars(
                select(ProjectSubtype).where(ProjectSubtype.project_id == 2).order_by(ProjectSubtype.name)
            ).all()

        self.assertIsNotNone(window_instance)
        self.assertIsNotNone(material_rule)
        assert window_instance is not None
        assert material_rule is not None
        subtype_ids = {subtype.name: subtype.id for subtype in subtypes}

        update_general = self.client.put(
            f"/api/v1/projects/2/instances/{window_instance.id}/materials/{material_rule.id}",
            headers={"X-Spec-Sheets-User": "editor"},
            json={
                "mode": "general",
                "entries": [{"subtype_id": None, "quantity": 7.5, "assembly_quantity": 1.0}],
            },
        )
        self.assertEqual(update_general.status_code, 200)

        with self.session_factory() as session:
            general_rows = session.scalars(
                select(ProjectBomEntry)
                .where(ProjectBomEntry.instance_id == window_instance.id, ProjectBomEntry.material_rule_id == material_rule.id)
            ).all()
        self.assertEqual(len(general_rows), 1)
        self.assertIsNone(general_rows[0].subtype_id)
        self.assertEqual(general_rows[0].quantity, 7.5)

        update_general_again = self.client.put(
            f"/api/v1/projects/2/instances/{window_instance.id}/materials/{material_rule.id}",
            headers={"X-Spec-Sheets-User": "editor"},
            json={
                "mode": "general",
                "entries": [{"subtype_id": None, "quantity": 15, "assembly_quantity": 5}],
            },
        )
        self.assertEqual(update_general_again.status_code, 200)

        with self.session_factory() as session:
            rewritten_general_rows = session.scalars(
                select(ProjectBomEntry)
                .where(ProjectBomEntry.instance_id == window_instance.id, ProjectBomEntry.material_rule_id == material_rule.id)
            ).all()
        self.assertEqual(len(rewritten_general_rows), 1)
        self.assertEqual(rewritten_general_rows[0].quantity, 15)
        self.assertEqual(rewritten_general_rows[0].assembly_quantity, 5)

        update_per_subtype = self.client.put(
            f"/api/v1/projects/2/instances/{window_instance.id}/materials/{material_rule.id}",
            headers={"X-Spec-Sheets-User": "editor"},
            json={
                "mode": "per_subtype",
                "entries": [
                    {"subtype_id": subtype_ids["Premium"], "quantity": 4.1, "assembly_quantity": 0},
                    {"subtype_id": subtype_ids["Standard"], "quantity": 3.2, "assembly_quantity": 0},
                ],
            },
        )
        self.assertEqual(update_per_subtype.status_code, 200)

        with self.session_factory() as session:
            subtype_rows = session.scalars(
                select(ProjectBomEntry)
                .where(ProjectBomEntry.instance_id == window_instance.id, ProjectBomEntry.material_rule_id == material_rule.id)
                .order_by(ProjectBomEntry.subtype_id)
            ).all()
        self.assertEqual(len(subtype_rows), 2)
        self.assertTrue(all(row.subtype_id is not None for row in subtype_rows))

        detail = self.client.get("/api/v1/projects/2", headers={"X-Spec-Sheets-User": "editor"})
        self.assertEqual(detail.status_code, 200)
        windows = next(section for section in detail.json()["categories"] if section["name"] == "Windows")
        living_window = next(item for item in windows["instances"] if item["id"] == window_instance.id)
        glass_material = next(item for item in living_window["materials"] if item["rule_id"] == material_rule.id)
        self.assertEqual(glass_material["mode"], "per_subtype")
        self.assertEqual([row["subtype"] for row in glass_material["bom_entries"]], ["Standard", "Premium"])

    def test_comments_notifications_exports_and_approvals(self) -> None:
        comments_response = self.client.get("/api/v1/projects/2/comments", headers={"X-Spec-Sheets-User": "viewer"})
        self.assertEqual(comments_response.status_code, 200)
        self.assertGreaterEqual(len(comments_response.json()), 1)

        create_comment = self.client.post(
            "/api/v1/projects/2/comments",
            headers={"X-Spec-Sheets-User": "editor"},
            json={"instance_id": 2, "body": "Need wording update for @viewer before export."},
        )
        self.assertEqual(create_comment.status_code, 200)
        self.assertIn("viewer", create_comment.json()["mentions"])

        notifications = self.client.get("/api/v1/notifications", headers={"X-Spec-Sheets-User": "viewer"})
        self.assertEqual(notifications.status_code, 200)
        self.assertTrue(any(item["type"] == "comment_mention" for item in notifications.json()))

        create_export = self.client.post(
            "/api/v1/projects/2/exports",
            headers={"X-Spec-Sheets-User": "editor"},
            json={"kind": "commercial_pdf", "payload": {"include_accessories": True}},
        )
        self.assertEqual(create_export.status_code, 200)
        self.assertEqual(create_export.json()["kind"], "commercial_pdf")

        request_approval = self.client.post(
            "/api/v1/projects/2/approvals",
            headers={"X-Spec-Sheets-User": "editor"},
            json={"summary": "Approve commercial package wording."},
        )
        self.assertEqual(request_approval.status_code, 200)
        self.assertEqual(request_approval.json()["status"], "pending")

        approval_id = request_approval.json()["id"]
        decision = self.client.post(
            f"/api/v1/approvals/{approval_id}/decision",
            headers={"X-Spec-Sheets-User": "admin"},
            json={"status": "approved"},
        )
        self.assertEqual(decision.status_code, 200)
        self.assertEqual(decision.json()["status"], "approved")

    def test_materials_workbook_export_generates_downloadable_artifact(self) -> None:
        create_export = self.client.post(
            "/api/v1/projects/2/exports",
            headers={"X-Spec-Sheets-User": "editor"},
            json={"kind": "materials_workbook", "payload": {"group_by": "context"}},
        )
        self.assertEqual(create_export.status_code, 200)
        self.assertEqual(create_export.json()["kind"], "materials_workbook")
        self.assertEqual(create_export.json()["status"], "completed")
        artifact_uri = create_export.json()["artifact_uri"]
        self.assertTrue(artifact_uri.startswith("/exports/"))

        artifact_response = self.client.get(artifact_uri, headers={"X-Spec-Sheets-User": "viewer"})
        self.assertEqual(artifact_response.status_code, 200)

        workbook = load_workbook(filename=BytesIO(artifact_response.content))
        self.assertEqual(workbook.sheetnames, ["Total Materials", "By Context", "Assembly Kit"])

        totals = workbook["Total Materials"]
        totals_by_sku = {
            totals.cell(row=row_index, column=2).value: totals.cell(row=row_index, column=3).value
            for row_index in range(2, totals.max_row + 1)
            if totals.cell(row=row_index, column=2).value
        }
        self.assertEqual(totals_by_sku["MAT-001"], 24)
        self.assertEqual(totals_by_sku["MAT-002"], 3)
        self.assertAlmostEqual(float(totals_by_sku["MAT-003"]), 7.3, places=6)
        self.assertAlmostEqual(float(totals_by_sku["MAT-005"]), 5.25, places=6)
        self.assertIsNone(totals_by_sku["MAT-006"])

        context_values = [cell for row in workbook["By Context"].iter_rows(values_only=True) for cell in row if cell]
        self.assertIn("1.2. Windows", context_values)
        self.assertIn("Living Window", "".join(str(value) for value in context_values))
        self.assertIn("Laminated Glass Panel", context_values)

        assembly_values = [cell for row in workbook["Assembly Kit"].iter_rows(values_only=True) for cell in row if cell]
        self.assertIn("Anchor Screw 5x70", assembly_values)
        self.assertIn("Smart Lock Kit", assembly_values)

    def test_cost_model_workbook_export_requires_cost_model_permission(self) -> None:
        denied = self.client.post(
            "/api/v1/projects/2/exports",
            headers={"X-Spec-Sheets-User": "editor"},
            json={"kind": "cost_model_workbook", "payload": {}},
        )
        self.assertEqual(denied.status_code, 403)

    def test_cost_model_workbook_export_generates_restricted_formula_workbook(self) -> None:
        create_export = self.client.post(
            "/api/v1/projects/2/exports",
            headers={"X-Spec-Sheets-User": "ot"},
            json={"kind": "cost_model_workbook", "payload": {}},
        )
        self.assertEqual(create_export.status_code, 200)
        self.assertEqual(create_export.json()["kind"], "cost_model_workbook")
        self.assertEqual(create_export.json()["status"], "completed")
        artifact_uri = create_export.json()["artifact_uri"]
        self.assertTrue(artifact_uri.endswith(".xlsx"))

        denied_artifact = self.client.get(artifact_uri, headers={"X-Spec-Sheets-User": "viewer"})
        self.assertEqual(denied_artifact.status_code, 403)

        artifact_response = self.client.get(artifact_uri, headers={"X-Spec-Sheets-User": "ot"})
        self.assertEqual(artifact_response.status_code, 200)

        workbook = load_workbook(filename=BytesIO(artifact_response.content), data_only=False)
        self.assertEqual(workbook.sheetnames, ["By Instance", "Total Materials"])

        by_instance = workbook["By Instance"]
        self.assertEqual(
            [by_instance.cell(row=1, column=column_index).value for column_index in range(1, 10)],
            ["Instance", "Category", "Subtype", "Material", "SKU", "Unit", "Quantity", "Unit Price", "Cost"],
        )

        auxiliary_rows = [
            row
            for row in by_instance.iter_rows(min_row=2, values_only=True)
            if row[4] in {"AUX-001", "AUX-002"}
        ]
        self.assertEqual(len(auxiliary_rows), 2)
        aux_by_sku = {row[4]: row for row in auxiliary_rows}
        self.assertEqual(aux_by_sku["AUX-001"][0], "Auxiliary Materials")
        self.assertEqual(aux_by_sku["AUX-001"][6], 1)
        self.assertEqual(aux_by_sku["AUX-001"][7], 185000)
        self.assertEqual(aux_by_sku["AUX-002"][7], 32000)

        anchor_row_index = next(
            row_index
            for row_index in range(2, by_instance.max_row + 1)
            if by_instance.cell(row=row_index, column=5).value == "MAT-001"
        )
        self.assertEqual(by_instance.cell(row=anchor_row_index, column=9).value, f'=IF(OR(G{anchor_row_index}="",H{anchor_row_index}=""),"",G{anchor_row_index}*H{anchor_row_index})')
        by_instance_total_general_row = next(
            row_index
            for row_index in range(2, by_instance.max_row + 1)
            if by_instance.cell(row=row_index, column=4).value == "Total General"
        )
        self.assertEqual(
            by_instance.cell(row=by_instance_total_general_row, column=9).value,
            f'=SUMIFS(I2:I{by_instance_total_general_row - 1},C2:C{by_instance_total_general_row - 1},"General")',
        )

        totals = workbook["Total Materials"]
        self.assertEqual(
            [totals.cell(row=1, column=column_index).value for column_index in range(1, 8)],
            ["SKU", "Material", "Subtype", "Unit", "Quantity", "Unit Price", "Cost"],
        )
        total_anchor_row_index = next(
            row_index
            for row_index in range(2, totals.max_row + 1)
            if totals.cell(row=row_index, column=1).value == "MAT-001"
        )
        self.assertIn("SUMIFS('By Instance'!$G:$G", totals.cell(row=total_anchor_row_index, column=5).value)
        self.assertEqual(
            totals.cell(row=total_anchor_row_index, column=6).value,
            f'=IF(OR(E{total_anchor_row_index}="",E{total_anchor_row_index}=0,G{total_anchor_row_index}=""),"",G{total_anchor_row_index}/E{total_anchor_row_index})',
        )
        self.assertIn("SUMIFS('By Instance'!$I:$I", totals.cell(row=total_anchor_row_index, column=7).value)
        total_labels = [totals.cell(row=row_index, column=3).value for row_index in range(2, totals.max_row + 1)]
        self.assertIn("Total General", total_labels)
        totals_total_general_row = next(
            row_index
            for row_index in range(2, totals.max_row + 1)
            if totals.cell(row=row_index, column=3).value == "Total General"
        )
        self.assertEqual(
            totals.cell(row=totals_total_general_row, column=7).value,
            f'=SUMIFS(G2:G{totals_total_general_row - 1},C2:C{totals_total_general_row - 1},"General")',
        )

    def test_cost_model_view_returns_consolidated_rows_with_aggregate_and_per_subtype_adjustments(self) -> None:
        view_response = self.client.get(
            "/api/v1/projects/2/cost-model",
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(view_response.status_code, 200)
        view = view_response.json()
        self.assertEqual(view["project"]["id"], 2)

        rows_by_sku = {row["sku"]: row for row in view["rows"]}
        self.assertIn("MAT-001", rows_by_sku)
        anchor = rows_by_sku["MAT-001"]
        self.assertAlmostEqual(anchor["estimated_total_quantity"], 24.0, places=6)
        self.assertFalse(anchor["is_auxiliary"])
        self.assertEqual(anchor["adjustments"], [])

        aux_rows = [row for row in view["rows"] if row["is_auxiliary"]]
        self.assertGreaterEqual(len(aux_rows), 1)

        material_id = anchor["material_id"]
        upsert_aggregate = self.client.put(
            "/api/v1/projects/2/cost-model/adjustments",
            headers={"X-Spec-Sheets-User": "editor"},
            json={
                "material_id": material_id,
                "subtype_id": None,
                "adjusted_quantity": 30.5,
                "source_kind": "historic_consumption",
                "source_note": "Based on last quarter",
                "source_sample_houses": 12,
                "source_total_consumption": 366.0,
            },
        )
        self.assertEqual(upsert_aggregate.status_code, 200)
        aggregate_row = next(row for row in upsert_aggregate.json()["rows"] if row["sku"] == "MAT-001")
        self.assertEqual(len(aggregate_row["adjustments"]), 1)
        aggregate_adjustment = aggregate_row["adjustments"][0]
        self.assertIsNone(aggregate_adjustment["subtype_id"])
        self.assertAlmostEqual(aggregate_adjustment["adjusted_quantity"], 30.5, places=6)
        self.assertEqual(aggregate_adjustment["source_kind"], "historic_consumption")
        self.assertEqual(aggregate_adjustment["source_sample_houses"], 12)

        subtype_targets = [
            entry["subtype_id"]
            for entry in aggregate_row["subtypes"]
            if entry["subtype_id"] is not None
        ]
        if subtype_targets:
            subtype_id = subtype_targets[0]
            upsert_subtype = self.client.put(
                "/api/v1/projects/2/cost-model/adjustments",
                headers={"X-Spec-Sheets-User": "editor"},
                json={
                    "material_id": material_id,
                    "subtype_id": subtype_id,
                    "adjusted_quantity": 8.25,
                    "source_kind": "manual",
                },
            )
            self.assertEqual(upsert_subtype.status_code, 200)
            updated_row = next(row for row in upsert_subtype.json()["rows"] if row["sku"] == "MAT-001")
            self.assertEqual(len(updated_row["adjustments"]), 2)
            self.assertIn(subtype_id, {item["subtype_id"] for item in updated_row["adjustments"]})

            replace_subtype = self.client.put(
                "/api/v1/projects/2/cost-model/adjustments",
                headers={"X-Spec-Sheets-User": "editor"},
                json={
                    "material_id": material_id,
                    "subtype_id": subtype_id,
                    "adjusted_quantity": 9.0,
                },
            )
            self.assertEqual(replace_subtype.status_code, 200)
            replaced_row = next(row for row in replace_subtype.json()["rows"] if row["sku"] == "MAT-001")
            self.assertEqual(len(replaced_row["adjustments"]), 2)
            replaced_subtype_entry = next(
                item for item in replaced_row["adjustments"] if item["subtype_id"] == subtype_id
            )
            self.assertAlmostEqual(replaced_subtype_entry["adjusted_quantity"], 9.0, places=6)

            delete_subtype = self.client.request(
                "DELETE",
                "/api/v1/projects/2/cost-model/adjustments",
                headers={"X-Spec-Sheets-User": "editor"},
                json={"material_id": material_id, "subtype_id": subtype_id},
            )
            self.assertEqual(delete_subtype.status_code, 200)
            after_delete = next(row for row in delete_subtype.json()["rows"] if row["sku"] == "MAT-001")
            self.assertEqual(len(after_delete["adjustments"]), 1)
            self.assertIsNone(after_delete["adjustments"][0]["subtype_id"])

        delete_aggregate = self.client.request(
            "DELETE",
            "/api/v1/projects/2/cost-model/adjustments",
            headers={"X-Spec-Sheets-User": "editor"},
            json={"material_id": material_id, "subtype_id": None},
        )
        self.assertEqual(delete_aggregate.status_code, 200)
        cleared_row = next(row for row in delete_aggregate.json()["rows"] if row["sku"] == "MAT-001")
        self.assertEqual(cleared_row["adjustments"], [])

    @patch("app.services.cost_model._get_average_prices_for_products_batch")
    @patch("app.services.cost_model._open_connection")
    @patch("app.services.cost_model.erp_search_available")
    def test_cost_model_view_falls_back_to_live_erp_prices_when_cache_is_missing(
        self,
        erp_search_available_mock,
        open_connection_mock,
        average_prices_mock,
    ) -> None:
        with self.session_factory() as session:
            cache = session.scalar(select(ErpMaterialCache).where(ErpMaterialCache.sku == "MAT-001"))
            self.assertIsNotNone(cache)
            cache.average_price = None
            session.commit()

        class _DummyConnection:
            def cursor(self):
                return object()

        class _DummyContextManager:
            def __enter__(self):
                return _DummyConnection()

            def __exit__(self, exc_type, exc, tb):
                return False

        erp_search_available_mock.return_value = True
        open_connection_mock.return_value = _DummyContextManager()
        average_prices_mock.return_value = {"MAT-001": 4321.0}

        view_response = self.client.get(
            "/api/v1/projects/2/cost-model",
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(view_response.status_code, 200)

        rows_by_sku = {row["sku"]: row for row in view_response.json()["rows"]}
        self.assertEqual(rows_by_sku["MAT-001"]["price"], 4321.0)
        self.assertEqual(average_prices_mock.call_count, 1)

    def test_cost_model_view_endpoint_denies_viewer_edits_but_allows_read(self) -> None:
        view_response = self.client.get(
            "/api/v1/projects/2/cost-model",
            headers={"X-Spec-Sheets-User": "viewer"},
        )
        self.assertEqual(view_response.status_code, 200)

        denied = self.client.put(
            "/api/v1/projects/2/cost-model/adjustments",
            headers={"X-Spec-Sheets-User": "viewer"},
            json={"material_id": 1, "adjusted_quantity": 1.0},
        )
        self.assertEqual(denied.status_code, 403)

    @unittest.skipUnless(importlib.util.find_spec("reportlab"), "reportlab is not installed")
    def test_commercial_pdf_export_generates_browser_viewable_artifact(self) -> None:
        create_export = self.client.post(
            "/api/v1/projects/2/exports",
            headers={"X-Spec-Sheets-User": "editor"},
            json={"kind": "commercial_pdf", "payload": {}},
        )
        self.assertEqual(create_export.status_code, 200)
        self.assertEqual(create_export.json()["kind"], "commercial_pdf")
        self.assertEqual(create_export.json()["status"], "completed")
        artifact_uri = create_export.json()["artifact_uri"]
        self.assertTrue(artifact_uri.endswith(".pdf"))

        artifact_response = self.client.get(artifact_uri, headers={"X-Spec-Sheets-User": "viewer"})
        self.assertEqual(artifact_response.status_code, 200)
        self.assertEqual(artifact_response.headers.get("content-type"), "application/pdf")
        self.assertIn("inline", artifact_response.headers.get("content-disposition", ""))
        self.assertTrue(artifact_response.content.startswith(b"%PDF"))

    @unittest.skipUnless(importlib.util.find_spec("reportlab"), "reportlab is not installed")
    def test_full_technical_pdf_export_generates_browser_viewable_artifact(self) -> None:
        create_export = self.client.post(
            "/api/v1/projects/2/exports",
            headers={"X-Spec-Sheets-User": "editor"},
            json={"kind": "full_technical_pdf", "payload": {}},
        )
        self.assertEqual(create_export.status_code, 200)
        self.assertEqual(create_export.json()["kind"], "full_technical_pdf")
        self.assertEqual(create_export.json()["status"], "completed")
        artifact_uri = create_export.json()["artifact_uri"]
        self.assertTrue(artifact_uri.endswith(".pdf"))

        artifact_response = self.client.get(artifact_uri, headers={"X-Spec-Sheets-User": "viewer"})
        self.assertEqual(artifact_response.status_code, 200)
        self.assertEqual(artifact_response.headers.get("content-type"), "application/pdf")
        self.assertIn("inline", artifact_response.headers.get("content-disposition", ""))
        self.assertTrue(artifact_response.content.startswith(b"%PDF"))

    @unittest.skipUnless(importlib.util.find_spec("reportlab"), "reportlab is not installed")
    def test_detailed_material_pdf_export_generates_browser_viewable_artifact(self) -> None:
        create_export = self.client.post(
            "/api/v1/projects/2/exports",
            headers={"X-Spec-Sheets-User": "editor"},
            json={"kind": "detailed_material_pdf", "payload": {}},
        )
        self.assertEqual(create_export.status_code, 200)
        self.assertEqual(create_export.json()["kind"], "detailed_material_pdf")
        self.assertEqual(create_export.json()["status"], "completed")
        artifact_uri = create_export.json()["artifact_uri"]
        self.assertTrue(artifact_uri.endswith(".pdf"))

        artifact_response = self.client.get(artifact_uri, headers={"X-Spec-Sheets-User": "viewer"})
        self.assertEqual(artifact_response.status_code, 200)
        self.assertEqual(artifact_response.headers.get("content-type"), "application/pdf")
        self.assertIn("inline", artifact_response.headers.get("content-disposition", ""))
        self.assertTrue(artifact_response.content.startswith(b"%PDF"))

    def test_project_activity_groups_reuse_mutation_batch_id(self) -> None:
        detail = self.client.get("/api/v1/projects/2", headers={"X-Spec-Sheets-User": "editor"})
        self.assertEqual(detail.status_code, 200)
        windows = next(section for section in detail.json()["categories"] if section["name"] == "Windows")
        instance = windows["instances"][0]
        material = instance["materials"][0]

        batch_id = "batch-group-window-edit"
        headers = {
            "X-Spec-Sheets-User": "editor",
            "X-Mutation-Batch-Id": batch_id,
        }

        update_instance = self.client.put(
            f"/api/v1/projects/2/instances/{instance['id']}",
            headers=headers,
            json={
                "name": instance["name"],
                "short_name": instance["short_name"],
                "description": f"{instance['description']} Batch update",
                "short_description": instance["short_description"],
                "installation": instance["installation"],
                "unit_amount": instance["unit_amount"],
                "attribute_values": [
                    {"name": attribute["name"], "value": attribute["value"]}
                    for attribute in instance["editable_attributes"]
                ],
            },
        )
        self.assertEqual(update_instance.status_code, 200)

        updated_quantity = (material["bom_entries"][0]["quantity"] or 0) + 1
        update_material = self.client.put(
            f"/api/v1/projects/2/instances/{instance['id']}/materials/{material['rule_id']}",
            headers=headers,
            json={
                "mode": material["mode"],
                "entries": [
                    {
                        "subtype_id": entry["subtype_id"],
                        "quantity": updated_quantity if index == 0 else entry["quantity"],
                        "assembly_quantity": entry["assembly_quantity"],
                    }
                    for index, entry in enumerate(material["bom_entries"])
                ],
            },
        )
        self.assertEqual(update_material.status_code, 200)

        follow_up_update = self.client.put(
            f"/api/v1/projects/2/instances/{instance['id']}",
            headers={"X-Spec-Sheets-User": "editor"},
            json={
                "name": instance["name"],
                "short_name": instance["short_name"],
                "description": f"{instance['description']} Batch update",
                "short_description": f"{instance['short_description']} Solo update",
                "installation": instance["installation"],
                "unit_amount": instance["unit_amount"],
                "attribute_values": [
                    {"name": attribute["name"], "value": attribute["value"]}
                    for attribute in instance["editable_attributes"]
                ],
            },
        )
        self.assertEqual(follow_up_update.status_code, 200)

        activity = self.client.get("/api/v1/projects/2/activity", headers={"X-Spec-Sheets-User": "viewer"})
        self.assertEqual(activity.status_code, 200)

        grouped = activity.json()
        matching_group = next(group for group in grouped if group["title"] == f"Updated {instance['name']}" and group["entry_count"] == 2)
        self.assertEqual(len(matching_group["entries"]), 2)
        self.assertEqual(
            {entry["headline"] for entry in matching_group["entries"]},
            {"Component details changed", "Material quantities changed"},
        )
        self.assertTrue(all("Project" not in entry["headline"] for entry in matching_group["entries"]))
        material_entry = next(entry for entry in matching_group["entries"] if entry["headline"] == "Material quantities changed")
        self.assertTrue(any(change["label"].startswith("Standard") for change in material_entry["changes"]))
        detail_entry = next(entry for entry in matching_group["entries"] if entry["headline"] == "Component details changed")
        self.assertTrue(any(change["label"] == "Description" for change in detail_entry["changes"]))
        self.assertTrue(any(group["title"] == f"Updated {instance['name']}" and group["entry_count"] == 1 for group in grouped))

        history = self.client.get("/api/v1/activity", headers={"X-Spec-Sheets-User": "viewer"})
        self.assertEqual(history.status_code, 200)
        self.assertTrue(all("project" in group for group in history.json()))
        self.assertTrue(all(group["project"]["status"] == "execution" for group in history.json()))

    def test_material_changes_within_a_minute_are_collapsed_into_one_summary_group(self) -> None:
        detail = self.client.get("/api/v1/projects/2", headers={"X-Spec-Sheets-User": "editor"})
        self.assertEqual(detail.status_code, 200)
        windows = next(section for section in detail.json()["categories"] if section["name"] == "Windows")
        instance = windows["instances"][0]
        material = instance["materials"][0]

        first_row = material["bom_entries"][0]
        first_update = self.client.put(
            f"/api/v1/projects/2/instances/{instance['id']}/materials/{material['rule_id']}",
            headers={"X-Spec-Sheets-User": "editor"},
            json={
                "mode": material["mode"],
                "entries": [
                    {
                        "subtype_id": entry["subtype_id"],
                        "quantity": (entry["quantity"] or 0) + 2 if index == 0 else entry["quantity"],
                        "assembly_quantity": entry["assembly_quantity"],
                    }
                    for index, entry in enumerate(material["bom_entries"])
                ],
            },
        )
        self.assertEqual(first_update.status_code, 200)

        second_update = self.client.put(
            f"/api/v1/projects/2/instances/{instance['id']}/materials/{material['rule_id']}",
            headers={"X-Spec-Sheets-User": "editor"},
            json={
                "mode": material["mode"],
                "entries": [
                    {
                        "subtype_id": entry["subtype_id"],
                        "quantity": (entry["quantity"] or 0) + 2 if index == 0 else entry["quantity"],
                        "assembly_quantity": (entry["assembly_quantity"] or 0) + 3 if index == 0 else entry["assembly_quantity"],
                    }
                    for index, entry in enumerate(material["bom_entries"])
                ],
            },
        )
        self.assertEqual(second_update.status_code, 200)

        activity = self.client.get("/api/v1/projects/2/activity", headers={"X-Spec-Sheets-User": "viewer"})
        self.assertEqual(activity.status_code, 200)
        grouped = activity.json()

        merged_group = next(
            group
            for group in grouped
            if group["title"] == f"Updated materials for {instance['name']}" and group["actor"] == "Project Editor"
        )
        self.assertEqual(merged_group["entry_count"], 1)
        self.assertEqual(len(merged_group["entries"]), 1)

        merged_entry = merged_group["entries"][0]
        self.assertEqual(merged_entry["headline"], "Material quantities changed")
        self.assertEqual(merged_entry["subject_name"], material["material_name"])
        changes_by_label = {change["label"]: (change["before"], change["after"]) for change in merged_entry["changes"]}
        self.assertEqual(
            changes_by_label.get(f"{first_row['subtype']} quantity"),
            (str(first_row["quantity"]), str((first_row["quantity"] or 0) + 2)),
        )
        self.assertEqual(
            changes_by_label.get(f"{first_row['subtype']} assembly quantity"),
            (str(first_row["assembly_quantity"]), str((first_row["assembly_quantity"] or 0) + 3)),
        )

    def test_legacy_deleted_item_activity_recovers_subject_name_from_project_instances(self) -> None:
        with self.session_factory() as session:
            project = session.scalar(select(Project).where(Project.id == 2))
            self.assertIsNotNone(project)
            component = session.scalar(
                select(CatalogComponent).where(CatalogComponent.component_type == ComponentType.ITEM).order_by(CatalogComponent.id)
            )
            self.assertIsNotNone(component)
            assert project is not None
            assert component is not None

            session.add_all(
                [
                    ProjectInstance(
                        project=project,
                        component=component,
                        category_id=component.category_id,
                        instance_type=component.component_type,
                        name="Bedroom Door",
                        description="Interior MDF door with bottom undercut for bedrooms.",
                        installation="Standard bedroom installation.",
                    ),
                    ProjectInstance(
                        project=project,
                        component=component,
                        category_id=component.category_id,
                        instance_type=component.component_type,
                        name="Bathroom Door",
                        description="Interior MDF door with vented lower grille for bathrooms and kitchens.",
                        installation="Bathroom and kitchen installation with vented opening.",
                    ),
                ]
            )

            group = ProjectActivityGroup(
                project=project,
                title="Items removed",
            )
            session.add(group)
            session.flush()

            details = build_activity_details(
                headline="Item removed",
                kind="item",
                changes=[
                    {"label": "Color", "before": "Blanco", "after": None},
                    {"label": "Recinto", "before": "Baño", "after": None},
                ],
            )
            details["legacy_details"] = (
                "Descripción: Interior MDF door with vented lower grille for bathrooms and kitchens.\n"
                "Instalación: Bathroom and kitchen installation with vented opening.\n"
                "Atributos eliminados:\n"
                "  Recinto: Baño\n"
                "Materiales eliminados:\n"
                "  ADHESIVO DE CONTACTO AGOREX 60: Cantidad 0.03, Kit 0.0"
            )
            session.add(
                ProjectActivityLog(
                    project=project,
                    group=group,
                    entity_type="ItemInstance",
                    action="legacy_deleted",
                    details=details,
                )
            )
            session.commit()

        activity = self.client.get("/api/v1/projects/2/activity", headers={"X-Spec-Sheets-User": "viewer"})
        self.assertEqual(activity.status_code, 200)
        grouped = activity.json()

        removed_group = next(group for group in grouped if group["title"] == "Items removed")
        removed_entry = next(entry for entry in removed_group["entries"] if entry["headline"] == "Item removed")
        self.assertEqual(removed_entry["subject_name"], "Bathroom Door")

    def test_v1_catalog_and_project_instance_requests_preserve_short_description(self) -> None:
        create_component_response = self.client.post(
            "/api/v1/catalog/components",
            headers={"X-Spec-Sheets-User": "editor"},
            json={
                "category_id": 6,
                "component_type": "item",
                "name": "Vanity Module",
                "short_name": "VAN-01",
                "description": "Bathroom vanity module",
                "short_description": "Commercial vanity",
                "installation": "Install against finished wall.",
                "unit_type": "unit",
            },
        )
        self.assertEqual(create_component_response.status_code, 200)
        component_id = create_component_response.json()["component_id"]

        catalog = self.client.get("/api/v1/catalog?category_id=6", headers={"X-Spec-Sheets-User": "editor"})
        self.assertEqual(catalog.status_code, 200)
        created_component = next(item for item in catalog.json()["selected"]["components"] if item["id"] == component_id)
        self.assertEqual(created_component["short_description"], "Commercial vanity")

        create_instance_response = self.client.post(
            "/api/v1/projects/2/instances",
            headers={"X-Spec-Sheets-User": "editor"},
            json={
                "category_id": 6,
                "component_id": component_id,
                "name": "Vanity Instance A",
                "short_name": "VIA-01",
                "description": "Bathroom vanity instance",
                "short_description": "Client package vanity",
                "installation": "Install in bath 01.",
                "unit_amount": 1,
                "attribute_values": [],
            },
        )
        self.assertEqual(create_instance_response.status_code, 200)
        instance_id = create_instance_response.json()["instance_id"]

        project_detail = self.client.get("/api/v1/projects/2", headers={"X-Spec-Sheets-User": "editor"})
        self.assertEqual(project_detail.status_code, 200)
        kitchens = next(section for section in project_detail.json()["categories"] if section["name"] == "Kitchens")
        created_instance = next(item for item in kitchens["instances"] if item["id"] == instance_id)
        self.assertEqual(created_instance["short_description"], "Client package vanity")

    def test_dashboard_and_public_api_surfaces(self) -> None:
        dashboard = self.client.get(
            "/api/v1/dashboard/projects/2/materials",
            headers={"X-Spec-Sheets-User": "admin"},
        )
        self.assertEqual(dashboard.status_code, 200)
        row = next(item for item in dashboard.json()["rows"] if item["sku"] == "MAT-003")
        self.assertEqual(row["material_name"], "Laminated Glass Panel")
        self.assertGreaterEqual(row["shortage"], 0)

        public_projects = self.client.get("/api/v1/public/projects")
        self.assertEqual(public_projects.status_code, 200)
        public_names = [project["name"] for project in public_projects.json()["projects"]]
        self.assertIn("Casa Robles - Block A", public_names)
        self.assertIn("Casa Alerce - Delivered", public_names)

        public_skus = self.client.get("/api/v1/public/projects/2/skus")
        self.assertEqual(public_skus.status_code, 200)
        self.assertIn("MAT-006", public_skus.json()["skus"])

    @patch("app.main.get_material_dashboard_house_start_comparison")
    @patch("app.main.get_material_dashboard_history")
    @patch("app.main.get_material_dashboard_detail")
    @patch("app.main.get_material_dashboard_cost_centers")
    @patch("app.main.get_recent_material_dashboard")
    def test_material_dashboard_api_surfaces_and_permissions(
        self,
        recent_dashboard_mock,
        cost_centers_mock,
        detail_mock,
        movement_history_mock,
        house_comparison_mock,
    ) -> None:
        recent_dashboard_mock.return_value = {
            "materials": [
                {
                    "sku": "ERP-001",
                    "material_name": "Steel Stud 90",
                    "unit": "UN",
                    "last_movement_date": "2026-03-10",
                    "movement_quantity_60d": 180.0,
                    "movement_count_60d": 6,
                }
            ],
            "movement_window_days": 60,
            "ceco_filters": ["CC-01"],
            "generated_at": "2026-03-12T12:00:00",
        }
        cost_centers_mock.return_value = {
            "cecos": [
                {"code": "CC-01", "name": "Produccion"},
                {"code": "CC-02", "name": "Despacho"},
            ]
        }
        detail_mock.return_value = {
            "sku": "ERP-001",
            "material_name": "Steel Stud 90",
            "unit": "UN",
            "movement_quantity_30d": 75.0,
            "stock_on_hand": 32.0,
            "pending_purchase_quantity": 48.0,
            "average_price": 2500.0,
            "average_lead_time_days": 9.5,
            "median_lead_time_days": 9.0,
            "max_lead_time_days": 12.0,
            "lead_time_sample_count": 4,
            "average_daily_outgoing_30d": 2.5,
            "days_of_stock_30d": 12.8,
            "reorder_date_recent_rate": "2026-03-14",
            "last_purchase_order": {
                "date": "2026-03-01",
                "number": "OC-123",
                "estimated_delivery": "2026-03-18",
            },
            "generated_at": "2026-03-12T12:03:00",
        }
        movement_history_mock.return_value = {
            "sku": "ERP-001",
            "movement_days": 90,
            "ceco_filters": ["CC-01"],
            "range_start": "2025-12-13",
            "range_end": "2026-03-12",
            "movements": [
                {"date": "2026-03-10", "quantity": 10.0},
                {"date": "2026-03-11", "quantity": 5.0},
            ],
            "generated_at": "2026-03-12T12:05:00",
        }
        house_comparison_mock.return_value = {
            "sku": "ERP-001",
            "house_type_id": 3,
            "house_type_name": "Casa Base",
            "number_of_modules": 2,
            "movement_days": 12,
            "ceco_filters": ["CC-01"],
            "range_start": "2026-03-01",
            "range_end": "2026-03-12",
            "total_material_quantity": 45.0,
            "total_house_starts": 9,
            "material_per_house": 5.0,
            "latest_house_start_date": "2026-03-12",
            "points": [
                {
                    "date": "2026-03-10",
                    "material_quantity": 10.0,
                    "house_starts": 2,
                    "cumulative_material_quantity": 10.0,
                    "cumulative_house_starts": 2,
                    "material_per_house": 5.0,
                }
            ],
            "generated_at": "2026-03-12T12:07:00",
        }

        dashboard = self.client.get(
            "/api/v1/dashboard/materials?ceco=CC-01",
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(dashboard.status_code, 200)
        self.assertEqual(dashboard.json()["materials"][0]["sku"], "ERP-001")
        recent_dashboard_mock.assert_called_once()

        cecos = self.client.get(
            "/api/v1/dashboard/materials/cecos",
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(cecos.status_code, 200)
        self.assertEqual(cecos.json()["cecos"][0]["code"], "CC-01")

        detail = self.client.get(
            "/api/v1/dashboard/materials/ERP-001?ceco=CC-01",
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(detail.status_code, 200)
        self.assertEqual(detail.json()["stock_on_hand"], 32.0)

        dashboard_post = self.client.post(
            "/api/v1/dashboard/materials",
            json={"excluded_cecos": ["CC-02"]},
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(dashboard_post.status_code, 200)
        self.assertEqual(dashboard_post.json()["materials"][0]["sku"], "ERP-001")
        self.assertEqual(recent_dashboard_mock.call_args.kwargs["cost_centers"], [])
        self.assertEqual(recent_dashboard_mock.call_args.kwargs["excluded_cost_centers"], ["CC-02"])

        detail_post = self.client.post(
            "/api/v1/dashboard/materials/ERP-001",
            json={"excluded_cecos": ["CC-02"]},
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(detail_post.status_code, 200)
        self.assertEqual(detail_post.json()["stock_on_hand"], 32.0)
        self.assertEqual(detail_mock.call_args.kwargs["cost_centers"], [])
        self.assertEqual(detail_mock.call_args.kwargs["excluded_cost_centers"], ["CC-02"])

        movements = self.client.get(
            "/api/v1/dashboard/materials/ERP-001/movements?ceco=CC-01",
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(movements.status_code, 200)
        self.assertEqual(movements.json()["movements"][0]["quantity"], 10.0)

        movements_post = self.client.post(
            "/api/v1/dashboard/materials/ERP-001/movements",
            json={"excluded_cecos": ["CC-02"], "start_date": "2026-03-01", "end_date": "2026-03-12"},
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(movements_post.status_code, 200)
        self.assertEqual(movements_post.json()["movements"][0]["quantity"], 10.0)
        self.assertEqual(movement_history_mock.call_args.kwargs["cost_centers"], [])
        self.assertEqual(movement_history_mock.call_args.kwargs["excluded_cost_centers"], ["CC-02"])

        house_comparison = self.client.post(
            "/api/v1/dashboard/materials/ERP-001/house-comparison",
            json={"excluded_cecos": ["CC-02"], "house_type_id": 3, "start_date": "2026-03-01", "end_date": "2026-03-12"},
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(house_comparison.status_code, 200)
        self.assertEqual(house_comparison.json()["house_type_id"], 3)
        self.assertEqual(house_comparison_mock.call_args.kwargs["cost_centers"], [])

        denied = self.client.get(
            "/api/v1/dashboard/materials",
            headers={"X-Spec-Sheets-User": "viewer"},
        )
        self.assertEqual(denied.status_code, 403)

    @patch("app.main.require_project_view")
    @patch("app.main.get_material_dashboard_project_quantity_map")
    @patch("app.main.get_material_dashboard_economic_metrics")
    def test_material_dashboard_economic_metrics_api(
        self,
        economic_metrics_mock,
        project_quantity_map_mock,
        require_project_view_mock,
    ) -> None:
        economic_metrics_mock.return_value = {
            "house_type_id": 3,
            "project_id": 2,
            "ceco_filters": [],
            "range_start": "2026-03-01",
            "range_end": "2026-03-12",
            "total_house_starts": 9,
            "metrics": [
                {
                    "sku": "ERP-001",
                    "material_per_house": 5.0,
                    "predicted_quantity_per_house": 4.0,
                    "consumption_delta_percent": 25.0,
                    "consumption_cost_delta_per_house": 2500.0,
                    "average_price": 2500.0,
                }
            ],
            "generated_at": "2026-03-12T12:07:00",
        }
        project_quantity_map_mock.return_value = (SimpleNamespace(id=2), {"ERP-001": 4.0})

        response = self.client.post(
            "/api/v1/dashboard/materials/economic-metrics",
            json={"excluded_cecos": ["CC-02"], "house_type_id": 3, "project_id": 2, "start_date": "2026-03-01", "end_date": "2026-03-12"},
            headers={"X-Spec-Sheets-User": "editor"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["metrics"][0]["consumption_delta_percent"], 25.0)
        require_project_view_mock.assert_called_once()
        self.assertEqual(project_quantity_map_mock.call_args.kwargs["project_id"], 2)
        self.assertEqual(economic_metrics_mock.call_args.kwargs["project_quantity_by_sku"], {"ERP-001": 4.0})

        denied = self.client.post(
            "/api/v1/dashboard/materials/economic-metrics",
            json={"house_type_id": 3},
            headers={"X-Spec-Sheets-User": "viewer"},
        )
        self.assertEqual(denied.status_code, 403)

    @patch("app.services.dashboard.get_recent_movement_materials")
    def test_material_dashboard_server_cache_reuses_recent_dashboard_payload(self, recent_movement_mock) -> None:
        recent_movement_mock.return_value = [
            {
                "sku": "ERP-001",
                "material_name": "Steel Stud 90",
                "unit": "UN",
                "last_movement_date": "2026-03-10",
                "movement_quantity_60d": 180.0,
                "movement_count_60d": 6,
            }
        ]

        with self.session_factory() as session:
            first = get_recent_material_dashboard(self.settings, session=session, cost_centers=["CC-01"])
            second = get_recent_material_dashboard(self.settings, session=session, cost_centers=["CC-01"])
            cached_entries = session.scalars(select(MaterialDashboardCacheEntry)).all()

        self.assertEqual(first["materials"][0]["sku"], "ERP-001")
        self.assertEqual(first, second)
        self.assertEqual(recent_movement_mock.call_count, 1)
        self.assertEqual(len(cached_entries), 1)

    @patch("app.services.dashboard.get_recent_movement_materials")
    def test_material_dashboard_server_cache_ignores_duplicate_insert_race(self, recent_movement_mock) -> None:
        recent_movement_mock.return_value = [
            {
                "sku": "ERP-001",
                "material_name": "Steel Stud 90",
                "unit": "UN",
                "last_movement_date": "2026-03-10",
                "movement_quantity_60d": 180.0,
                "movement_count_60d": 6,
            }
        ]

        with self.session_factory() as session:
            original_flush = session.flush
            inserted_competing_entry = False

            def flush_with_race(*args, **kwargs):
                nonlocal inserted_competing_entry
                if not inserted_competing_entry:
                    inserted_competing_entry = True
                    with self.session_factory() as competing_session:
                        get_recent_material_dashboard(
                            self.settings,
                            session=competing_session,
                            excluded_cost_centers=["CC-01"],
                        )
                        competing_session.commit()
                return original_flush(*args, **kwargs)

            session.flush = flush_with_race  # type: ignore[method-assign]
            result = get_recent_material_dashboard(
                self.settings,
                session=session,
                excluded_cost_centers=["CC-01"],
            )

        with self.session_factory() as verification_session:
            cached_entries = verification_session.scalars(select(MaterialDashboardCacheEntry)).all()

        self.assertEqual(result["materials"][0]["sku"], "ERP-001")
        self.assertEqual(len(cached_entries), 1)

    @patch("app.services.dashboard.get_recent_movement_materials")
    def test_material_dashboard_server_cache_returns_payload_when_cache_update_times_out(
        self,
        recent_movement_mock,
    ) -> None:
        recent_movement_mock.side_effect = [
            [
                {
                    "sku": "ERP-001",
                    "material_name": "Steel Stud 90",
                    "unit": "UN",
                    "last_movement_date": "2026-03-10",
                    "movement_quantity_60d": 180.0,
                    "movement_count_60d": 6,
                }
            ],
            [
                {
                    "sku": "ERP-002",
                    "material_name": "Track 90",
                    "unit": "UN",
                    "last_movement_date": "2026-03-11",
                    "movement_quantity_60d": 90.0,
                    "movement_count_60d": 4,
                }
            ],
        ]

        with self.session_factory() as seed_session:
            get_recent_material_dashboard(self.settings, session=seed_session, cost_centers=["CC-01"])
            entry = seed_session.scalar(
                select(MaterialDashboardCacheEntry).where(MaterialDashboardCacheEntry.cache_kind == "list")
            )
            self.assertIsNotNone(entry)
            entry.expires_at = datetime.now(timezone.utc) - timedelta(minutes=1)
            seed_session.commit()

        with self.session_factory() as session:
            def flush_with_timeout(*args, **kwargs):
                raise OperationalError(
                    "UPDATE material_dashboard_cache_entries",
                    {},
                    Exception(
                        'canceling statement due to statement timeout CONTEXT: while updating tuple (18,7) in relation "material_dashboard_cache_entries"'
                    ),
                )

            session.flush = flush_with_timeout  # type: ignore[method-assign]
            result = get_recent_material_dashboard(self.settings, session=session, cost_centers=["CC-01"])

        with self.session_factory() as verification_session:
            cached_entry = verification_session.scalar(
                select(MaterialDashboardCacheEntry).where(MaterialDashboardCacheEntry.cache_kind == "list")
            )

        self.assertEqual(result["materials"][0]["sku"], "ERP-002")
        self.assertEqual(recent_movement_mock.call_count, 2)
        self.assertIsNotNone(cached_entry)
        self.assertEqual(cached_entry.payload["materials"][0]["sku"], "ERP-001")

    @patch("app.services.dashboard.get_average_prices_for_products")
    @patch("app.services.dashboard.get_material_dashboard_house_start_summary")
    @patch("app.services.dashboard.get_recent_material_dashboard")
    def test_material_dashboard_economic_metrics_compute_and_cache(
        self,
        recent_dashboard_mock,
        house_start_summary_mock,
        average_prices_mock,
    ) -> None:
        recent_dashboard_mock.return_value = {
            "materials": [
                {
                    "sku": "ERP-001",
                    "material_name": "Steel Stud 90",
                    "unit": "UN",
                    "last_movement_date": "2026-03-10",
                    "movement_quantity_60d": 45.0,
                    "movement_count_60d": 6,
                },
                {
                    "sku": "ERP-002",
                    "material_name": "Track 90",
                    "unit": "UN",
                    "last_movement_date": "2026-03-09",
                    "movement_quantity_60d": 9.0,
                    "movement_count_60d": 3,
                },
            ],
            "movement_window_days": 12,
            "ceco_filters": ["CC-01"],
            "generated_at": "2026-03-12T12:00:00",
        }
        house_start_summary_mock.return_value = {
            "house_type_id": 3,
            "house_type_name": "Casa Base",
            "number_of_modules": 2,
            "movement_days": 12,
            "ceco_filters": ["CC-01"],
            "range_start": "2026-03-01",
            "range_end": "2026-03-12",
            "total_house_starts": 9,
            "latest_house_start_date": "2026-03-12",
            "generated_at": "2026-03-12T12:05:00",
        }
        average_prices_mock.return_value = {
            "ERP-001": 2500.0,
            "ERP-002": 1500.0,
        }

        with self.session_factory() as session:
            first = get_material_dashboard_economic_metrics(
                self.settings,
                session=session,
                house_type_id=3,
                project_id=2,
                project_quantity_by_sku={"ERP-001": 4.0, "ERP-002": 1.0},
                start_date=date(2026, 3, 1),
                end_date=date(2026, 3, 12),
                cost_centers=["CC-01"],
            )
            second = get_material_dashboard_economic_metrics(
                self.settings,
                session=session,
                house_type_id=3,
                project_id=2,
                project_quantity_by_sku={"ERP-001": 4.0, "ERP-002": 1.0},
                start_date=date(2026, 3, 1),
                end_date=date(2026, 3, 12),
                cost_centers=["CC-01"],
            )
            cached_entries = session.scalars(
                select(MaterialDashboardCacheEntry).where(MaterialDashboardCacheEntry.cache_kind == "economics")
            ).all()

        metrics_by_sku = {row["sku"]: row for row in first["metrics"]}
        self.assertEqual(first, second)
        self.assertEqual(metrics_by_sku["ERP-001"]["material_per_house"], 5.0)
        self.assertEqual(metrics_by_sku["ERP-001"]["consumption_delta_percent"], 25.0)
        self.assertEqual(metrics_by_sku["ERP-001"]["consumption_cost_delta_per_house"], 2500.0)
        self.assertEqual(metrics_by_sku["ERP-002"]["consumption_delta_percent"], 0.0)
        self.assertEqual(recent_dashboard_mock.call_count, 1)
        self.assertEqual(house_start_summary_mock.call_count, 1)
        self.assertEqual(average_prices_mock.call_count, 1)
        self.assertEqual(len(cached_entries), 1)

    @patch("app.services.dashboard.get_material_movement_details")
    @patch("app.services.dashboard.get_material_movement_history")
    def test_material_dashboard_history_cache_hashes_long_filter_keys(
        self,
        movement_history_mock,
        movement_details_mock,
    ) -> None:
        movement_history_mock.return_value = [{"date": "2026-03-18", "quantity": 10.0}]
        movement_details_mock.return_value = [{"date": "2026-03-18", "quantity": 10.0}]
        excluded_cost_centers = [f"{index:02d}-{index:02d}-{index:02d}" for index in range(1, 31)]

        with self.session_factory() as session:
            first = get_material_dashboard_history(
                self.settings,
                "ERP-001",
                session=session,
                excluded_cost_centers=excluded_cost_centers,
                history_days=90,
                start_date=date(2025, 12, 22),
                end_date=date(2026, 3, 19),
            )
            second = get_material_dashboard_history(
                self.settings,
                "ERP-001",
                session=session,
                excluded_cost_centers=excluded_cost_centers,
                history_days=90,
                start_date=date(2025, 12, 22),
                end_date=date(2026, 3, 19),
            )
            cached_entries = session.scalars(select(MaterialDashboardCacheEntry)).all()

        self.assertEqual(first, second)
        self.assertEqual(movement_history_mock.call_count, 1)
        self.assertEqual(movement_details_mock.call_count, 1)
        self.assertEqual(len(cached_entries), 1)
        self.assertLessEqual(len(cached_entries[0].cache_key), 255)

    def test_app_lifespan_requires_existing_schema(self) -> None:
        Base.metadata.drop_all(self.engine)
        settings = Settings(database_url=self.test_database_url, seed_demo_data=False, require_schema=True)
        app = create_app(settings)

        async def run_lifespan() -> None:
            async with app.router.lifespan_context(app):
                pass

        with self.assertRaises(RuntimeError):
            asyncio.run(run_lifespan())

    def test_app_lifespan_seeds_after_schema_exists(self) -> None:
        Base.metadata.drop_all(self.engine)
        Base.metadata.create_all(self.engine)

        settings = Settings(database_url=self.test_database_url, seed_demo_data=True, require_schema=True)
        app = create_app(settings)

        async def run_lifespan() -> None:
            async with app.router.lifespan_context(app):
                pass

        asyncio.run(run_lifespan())

        with self.session_factory() as session:
            self.assertIsNotNone(session.query(CatalogCategory).first())


class MaterialStudyGroupTests(ServiceLayerTests):
    def test_update_group_reuses_existing_members_for_unchanged_skus(self) -> None:
        with self.session_factory() as session:
            created = create_material_study_group(
                session,
                name="Insulation",
                description="Normalized in m2",
                study_unit="m2",
                members=[
                    {"sku": "INS-S", "material_name": "Insulation Small", "unit": "roll", "factor_to_study_unit": 2},
                    {"sku": "INS-L", "material_name": "Insulation Large", "unit": "roll", "factor_to_study_unit": 10},
                ],
            )
            session.commit()

        with self.session_factory() as session:
            updated = update_material_study_group(
                session,
                created["group_id"],
                name="Insulation",
                description="Normalized in m2",
                study_unit="m2",
                members=[
                    {
                        "sku": "INS-L",
                        "material_name": "Insulation Large Updated",
                        "unit": "panel",
                        "factor_to_study_unit": 12,
                    },
                    {
                        "sku": "INS-XL",
                        "material_name": "Insulation Extra Large",
                        "unit": "panel",
                        "factor_to_study_unit": 15,
                    },
                ],
            )
            session.commit()

            persisted_group = session.scalar(select(MaterialStudyGroup).where(MaterialStudyGroup.id == created["group_id"]))

        self.assertIsNotNone(updated)
        assert updated is not None
        self.assertEqual(updated["member_count"], 2)
        self.assertEqual([member["sku"] for member in updated["members"]], ["INS-L", "INS-XL"])
        self.assertEqual(updated["members"][0]["material_name"], "Insulation Large Updated")
        self.assertEqual(updated["members"][0]["unit"], "panel")
        self.assertEqual(updated["members"][0]["factor_to_study_unit"], 12.0)
        self.assertEqual(updated["members"][0]["display_order"], 0)
        self.assertEqual(updated["members"][1]["display_order"], 1)
        self.assertIsNotNone(persisted_group)
        assert persisted_group is not None
        self.assertEqual([member.sku for member in persisted_group.members], ["INS-L", "INS-XL"])

    @patch("app.services.material_groups.get_recent_movement_materials")
    def test_group_list_normalizes_recent_movement_metrics(self, recent_movement_mock) -> None:
        with self.session_factory() as session:
            create_material_study_group(
                session,
                name="Insulation",
                description="Normalized in m2",
                study_unit="m2",
                members=[
                    {"sku": "INS-S", "material_name": "Insulation Small", "unit": "roll", "factor_to_study_unit": 2},
                    {"sku": "INS-L", "material_name": "Insulation Large", "unit": "roll", "factor_to_study_unit": 10},
                ],
            )
            session.commit()

        recent_movement_mock.return_value = [
            {
                "sku": "INS-S",
                "material_name": "Insulation Small",
                "unit": "roll",
                "last_movement_date": "2026-03-10",
                "movement_quantity_60d": 3.0,
                "movement_count_60d": 2,
            },
            {
                "sku": "INS-L",
                "material_name": "Insulation Large",
                "unit": "roll",
                "last_movement_date": "2026-03-12",
                "movement_quantity_60d": 4.0,
                "movement_count_60d": 1,
            },
        ]

        with self.session_factory() as session:
            response = get_material_dashboard_groups(self.settings, session=session, movement_days=60)

        self.assertEqual(len(response["groups"]), 1)
        self.assertEqual(response["groups"][0]["movement_quantity_60d"], 46.0)
        self.assertEqual(response["groups"][0]["movement_count_60d"], 3)
        self.assertEqual(response["groups"][0]["last_movement_date"], "2026-03-12")
        self.assertEqual(response["groups"][0]["sku"], "GROUP:1")

    @patch("app.services.material_groups.get_material_movement_details")
    @patch("app.services.material_groups.get_material_movement_history")
    @patch("app.services.material_groups.get_material_procurement_details")
    def test_group_detail_and_history_normalize_member_values(
        self,
        procurement_mock,
        movement_history_mock,
        movement_details_mock,
    ) -> None:
        with self.session_factory() as session:
            create_material_study_group(
                session,
                name="Insulation",
                description="Normalized in m2",
                study_unit="m2",
                members=[
                    {"sku": "INS-S", "material_name": "Insulation Small", "unit": "roll", "factor_to_study_unit": 2},
                    {"sku": "INS-L", "material_name": "Insulation Large", "unit": "roll", "factor_to_study_unit": 10},
                ],
            )
            session.commit()

        procurement_by_sku = {
            "INS-S": {
                "material_name": "Insulation Small",
                "unit": "roll",
                "movement_quantity_30d": 5.0,
                "stock_on_hand": 8.0,
                "pending_purchase_quantity": 2.0,
            },
            "INS-L": {
                "material_name": "Insulation Large",
                "unit": "roll",
                "movement_quantity_30d": 2.0,
                "stock_on_hand": 3.0,
                "pending_purchase_quantity": 1.0,
            },
        }
        procurement_mock.side_effect = lambda _settings, sku, **_kwargs: procurement_by_sku[sku]

        movement_history_by_sku = {
            "INS-S": [
                {"date": "2026-03-10", "quantity": 1.0},
                {"date": "2026-03-11", "quantity": 2.0},
            ],
            "INS-L": [
                {"date": "2026-03-10", "quantity": 0.0},
                {"date": "2026-03-11", "quantity": 1.0},
            ],
        }
        movement_history_mock.side_effect = lambda _settings, sku, **_kwargs: movement_history_by_sku[sku]
        movement_details_mock.side_effect = lambda _settings, sku, **_kwargs: [
            {
                "date": "2026-03-11",
                "quantity": movement_history_by_sku[sku][-1]["quantity"],
                "ceco": "CC-01",
                "ceco_name": "Main",
                "desc_sub": "T2" if sku == "INS-S" else None,
                "movement_internal_number": f"MOV-{sku}",
                "line_count": 1,
            }
        ]

        with self.session_factory() as session:
            detail = get_material_dashboard_group_detail(self.settings, 1, session=session)
            history = get_material_dashboard_group_history(
                self.settings,
                1,
                session=session,
                start_date=date(2026, 3, 10),
                end_date=date(2026, 3, 11),
            )

        self.assertIsNotNone(detail)
        self.assertIsNotNone(history)
        assert detail is not None
        assert history is not None

        self.assertEqual(detail["movement_quantity_30d"], 30.0)
        self.assertEqual(detail["stock_on_hand"], 46.0)
        self.assertEqual(detail["pending_purchase_quantity"], 14.0)
        self.assertEqual(history["movements"][0]["quantity"], 2.0)
        self.assertEqual(history["movements"][1]["quantity"], 14.0)
        self.assertEqual(history["movement_details"][0]["sku"], "INS-L")
        self.assertEqual(history["movement_details"][0]["quantity"], 10.0)
        self.assertIsNone(history["movement_details"][0]["desc_sub"])
        self.assertEqual(history["movement_details"][1]["desc_sub"], "T2")


class MaterialDashboardBusinessDayTests(unittest.TestCase):
    @patch("app.services.dashboard.get_material_procurement_details")
    def test_detail_builder_uses_business_days_for_recent_rate_and_reorder_date(self, procurement_mock) -> None:
        class FixedDateTime(datetime):
            @classmethod
            def utcnow(cls):
                return cls(2026, 3, 16, 15, 0, 0)

        procurement_mock.return_value = {
            "material_name": "Steel Stud 90",
            "unit": "UN",
            "movement_quantity_30d": 84.0,
            "stock_on_hand": 210.0,
            "pending_purchase_quantity": 15.0,
            "average_price": 2500.0,
            "average_lead_time_days": 4.0,
            "median_lead_time_days": 4.0,
            "max_lead_time_days": 6.0,
            "lead_time_sample_count": 3,
            "last_purchase_order_date": "2026-03-01",
            "last_purchase_order_number": "OC-123",
            "last_purchase_order_estimated_delivery": "2026-03-18",
        }

        with patch("app.services.dashboard.datetime", FixedDateTime):
            detail = _build_material_dashboard_detail(Settings(), "ERP-001", cost_centers=["CC-01"])

        self.assertIsNotNone(detail)
        assert detail is not None

        today = FixedDateTime.utcnow().date()
        business_days = _count_business_days(today - timedelta(days=30), today)
        expected_average = round(84.0 / business_days, 2)
        expected_days_of_stock = round(210.0 / expected_average, 1)
        expected_reorder_offset = max(int(round(expected_days_of_stock - 6.0)), 0)

        self.assertEqual(detail["average_daily_outgoing_30d"], expected_average)
        self.assertEqual(detail["days_of_stock_30d"], expected_days_of_stock)
        self.assertEqual(detail["reorder_date_recent_rate"], _add_business_days(today, expected_reorder_offset).isoformat())

    def test_add_business_days_skips_weekends(self) -> None:
        self.assertEqual(_count_business_days(date(2026, 3, 13), date(2026, 3, 16)), 2)
        self.assertEqual(_add_business_days(date(2026, 3, 13), 1), date(2026, 3, 16))
        self.assertEqual(_add_business_days(date(2026, 3, 14), 0), date(2026, 3, 16))


class ErpLeadTimeSampleTests(unittest.TestCase):
    def test_last_purchase_orders_keep_approved_and_pending_states_only(self) -> None:
        row = SimpleNamespace(
            CodProd="ERP-001",
            fechaOC=date(2026, 4, 10),
            numoc="OC-123",
            FecFinalOC=None,
            CodEstado="PE",
            cantidadOrdenadaDetalle=10.0,
            cantidadIngresadaMovim=4.0,
            cantidadRecepcionNoInv=1.0,
        )

        class FakeCursor:
            def execute(self, sql, params):
                self.sql = sql
                self.params = params

            def fetchall(self):
                return [row]

        cursor = FakeCursor()

        result = _get_last_purchase_orders_for_products_batch(cursor, ["ERP-001"])

        self.assertIn("RTRIM(LTRIM(c.CodEstado)) IN (?,?)", cursor.sql)
        self.assertEqual(cursor.params[-2:], ["AP", "PE"])
        self.assertEqual(
            result["ERP-001"],
            (date(2026, 4, 10), "OC-123", 5.0, None, "PE"),
        )

    def test_calculate_delivery_time_stats_includes_median(self) -> None:
        with patch(
            "app.services.erp._get_lead_time_samples_for_product",
            return_value=[
                {"lead_time_days": 3},
                {"lead_time_days": 5},
                {"lead_time_days": 11},
                {"lead_time_days": 13},
            ],
        ):
            stats = _calculate_delivery_time_stats(object(), "ERP-001", limit=20)

        self.assertEqual(
            stats,
            {
                "average_lead_time_days": 8.0,
                "median_lead_time_days": 8.0,
                "max_lead_time_days": 13,
                "lead_time_sample_count": 4,
            },
        )

    def test_lead_time_sampling_scans_past_unusable_recent_orders(self) -> None:
        rows = [
            SimpleNamespace(
                fechaOC=f"2026-03-{day:02d}",
                FecFinalOC=None,
                numoc=f"OC-{day:03d}",
                OCNumInterOc=f"INT-{day:03d}",
                NumLinea="1",
                CodProd="ERP-001",
            )
            for day in range(15, 0, -1)
        ]

        class FakeCursor:
            def __init__(self, query_rows):
                self.query_rows = query_rows

            def execute(self, sql, params):
                self.sql = sql
                self.params = params

            def fetchall(self):
                return self.query_rows

        cursor = FakeCursor(rows)

        receipt_dates = [None, None, None, date(2026, 3, 12), date(2026, 3, 11), date(2026, 3, 10)]
        with patch("app.services.erp._fetch_first_receipt_date", side_effect=receipt_dates):
            samples = _get_lead_time_samples_for_product(cursor, "ERP-001", limit=3)

        self.assertEqual(len(samples), 3)
        self.assertTrue(all(sample["lead_time_days"] is not None for sample in samples))


if __name__ == "__main__":
    unittest.main()
